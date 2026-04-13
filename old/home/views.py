from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q, Count, Prefetch
from django.db import connection, transaction
from datetime import datetime, timedelta
import json
import os
import uuid
import time
import asyncio
import logging
from .models import TelegramAccount, Contact, Chat, Message, AIIntegration, AutoReplyRule, AutoReplyLog
from .utils import TelegramManager, import_contacts_from_excel, run_async, run_coroutine_in_thread
from django.db import close_old_connections, OperationalError
from django.utils import timezone
from django.conf import settings
from .telegram_monitor import start_monitoring, get_monitor

logger = logging.getLogger(__name__)

# Authentication Views
def login_view(request):
    """User login view"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            next_url = request.GET.get('next', 'dashboard')
            messages.success(request, f'Xush kelibsiz, {user.username}!')
            return redirect(next_url)
        else:
            messages.error(request, 'Login yoki parol noto\'g\'ri!')
    
    return render(request, 'home/login.html')

def register_view(request):
    """User registration view"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')
        
        # Validation
        if not username or not password:
            messages.error(request, 'Login va parol majburiy!')
            return render(request, 'home/register.html')
        
        if password != password_confirm:
            messages.error(request, 'Parollar mos kelmaydi!')
            return render(request, 'home/register.html')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Bu login band!')
            return render(request, 'home/register.html')
        
        if email and User.objects.filter(email=email).exists():
            messages.error(request, 'Bu email band!')
            return render(request, 'home/register.html')
        
        # Create user
        user = User.objects.create_user(username=username, email=email, password=password)
        login(request, user)
        messages.success(request, 'Ro\'yxatdan o\'tdingiz!')
        return redirect('dashboard')
    
    return render(request, 'home/register.html')

def logout_view(request):
    """User logout view"""
    logout(request)
    messages.success(request, 'Tizimdan chiqdingiz!')
    return redirect('login')

@login_required
def dashboard(request):
    """Main dashboard view"""
    # Start monitoring when user logs in
    try:
        monitor = get_monitor()
        if not monitor.running:
            start_monitoring()
    except Exception as e:
        logger.error(f"Failed to start monitor: {e}")
    
    telegram_accounts = TelegramAccount.objects.filter(user=request.user, is_active=True)
    
    # Statistics
    total_accounts = telegram_accounts.count()
    total_contacts = Contact.objects.filter(telegram_account__user=request.user).count()
    total_chats = Chat.objects.filter(telegram_account__user=request.user).count()
    total_messages = Message.objects.filter(telegram_account__user=request.user).count()
    
    context = {
        'telegram_accounts': telegram_accounts,
        'total_accounts': total_accounts,
        'total_contacts': total_contacts,
        'total_chats': total_chats,
        'total_messages': total_messages,
    }
    return render(request, 'home/dashboard.html', context)

@login_required
def telegram_accounts(request):
    """List all telegram accounts"""
    accounts = TelegramAccount.objects.filter(user=request.user).order_by('-created_at')
    
    context = {
        'accounts': accounts,
    }
    return render(request, 'home/telegram_accounts.html', context)

@login_required
def add_telegram_account(request):
    """Add new Telegram account"""
    if request.method == 'POST':
        api_id = request.POST.get('api_id')
        api_hash = request.POST.get('api_hash')
        phone_number = request.POST.get('phone_number')
        session_name = request.POST.get('session_name')
        
        if not all([api_id, api_hash, phone_number, session_name]):
            messages.error(request, 'Barcha maydonlar to\'ldirilishi kerak.')
            return render(request, 'home/add_telegram_account.html')
        
        # Check if session name is unique
        if TelegramAccount.objects.filter(session_name=session_name).exists():
            messages.error(request, 'Bu session nomi allaqachon mavjud.')
            return render(request, 'home/add_telegram_account.html')
        
        try:
            # Create telegram account
            telegram_account = TelegramAccount.objects.create(
                user=request.user,
                api_id=api_id,
                api_hash=api_hash,
                phone_number=phone_number,
                session_name=session_name,
                is_active=False  # Will be activated after verification
            )
            
            # Store account ID in session for verification process
            request.session['telegram_account_id'] = telegram_account.id
            
            messages.success(request, 'Telegram account yaratildi. Endi tasdiqlash kodini kiriting.')
            return redirect('verify_telegram_account')
            
        except Exception as e:
            messages.error(request, f'Xatolik: {str(e)}')
    
    return render(request, 'home/add_telegram_account.html')

@login_required
def verify_telegram_account(request):
    """Verify Telegram account with phone code"""
    # Check if account_id is provided in GET parameters (for reactivation)
    account_id = request.GET.get('account_id')
    mode = request.GET.get('mode', 'verify')  # 'verify' or 'simple'
    
    if account_id:
        # Reactivation flow
        try:
            telegram_account = TelegramAccount.objects.get(id=account_id, user=request.user)
            # Store in session for verification process
            request.session['telegram_account_id'] = telegram_account.id
        except TelegramAccount.DoesNotExist:
            messages.error(request, 'Telegram account topilmadi.')
            return redirect('telegram_accounts')
    else:
        # Normal flow - get from session
        telegram_account_id = request.session.get('telegram_account_id')
        if not telegram_account_id:
            messages.error(request, 'Telegram account topilmadi.')
            return redirect('add_telegram_account')
        
        telegram_account = get_object_or_404(TelegramAccount, id=telegram_account_id, user=request.user)
    
    # Handle simple reactivation (without Telegram API)
    if mode == 'simple' and request.method == 'GET':
        telegram_account.is_active = True
        telegram_account.save()
        
        messages.success(request, f'Telegram account {telegram_account.phone_number} oddiy usulda faollashtirildi!')
        return redirect('telegram_accounts')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        step = request.POST.get('step')
        
        if action == 'send_code' or step == 'send_code':
            try:
                manager = TelegramManager(telegram_account)
                
                async def send_code():
                    await manager.connect()
                    result = await manager.send_code(telegram_account.phone_number)
                    await manager.disconnect()
                    return result
                
                result = run_async(send_code())
                
                if result['success']:
                    request.session['phone_code_hash'] = result['phone_code_hash']
                    messages.success(request, 'Tasdiqlash kodi yuborildi.')
                    return render(request, 'home/verify_telegram_account.html', {
                        'telegram_account': telegram_account,
                        'step': 'code',
                        'phone_number': telegram_account.phone_number,
                        'phone_code_hash': result['phone_code_hash']
                    })
                else:
                    messages.error(request, f'Xatolik: {result["error"]}')
            except Exception as e:
                messages.error(request, f'Xatolik: {str(e)}')
        
        elif action == 'verify_code' or step == 'code':
            code = request.POST.get('verification_code')
            password = request.POST.get('password', '')
            phone_code_hash = request.POST.get('phone_code_hash') or request.session.get('phone_code_hash')
            
            if not code:
                messages.error(request, 'Tasdiqlash kodini kiriting.')
            else:
                try:
                    manager = TelegramManager(telegram_account)
                    
                    async def verify():
                        await manager.connect()
                        result = await manager.sign_in(
                            telegram_account.phone_number, 
                            code, 
                            phone_code_hash, 
                            password if password else None
                        )
                        if result['success']:
                            # Get user info
                            me = await manager.get_me()
                            await manager.disconnect()
                            return {'success': True, 'user': me}
                        await manager.disconnect()
                        return result
                    
                    result = run_async(verify())
                    
                    if result['success']:
                        telegram_account.is_active = True
                        telegram_account.save()
                        
                        # Clean session
                        if 'telegram_account_id' in request.session:
                            del request.session['telegram_account_id']
                        if 'phone_code_hash' in request.session:
                            del request.session['phone_code_hash']
                        
                        messages.success(request, f'Telegram account {telegram_account.phone_number} muvaffaqiyatli tasdiqlandi!')
                        return render(request, 'home/verify_telegram_account.html', {
                            'step': 'success',
                            'telegram_account': telegram_account
                        })
                    elif result.get('needs_password'):
                        messages.warning(request, 'Ikki bosqichli tasdiqlash yoqilgan. Parolni kiriting.')
                        return render(request, 'home/verify_telegram_account.html', {
                            'telegram_account': telegram_account,
                            'step': 'password'
                        })
                    else:
                        messages.error(request, f'Xatolik: {result["error"]}')
                except Exception as e:
                    messages.error(request, f'Xatolik: {str(e)}')
        
        elif step == 'password':
            password = request.POST.get('password', '')
            phone_code_hash = request.session.get('phone_code_hash')
            
            if not password:
                messages.error(request, 'Parolni kiriting.')
            else:
                try:
                    manager = TelegramManager(telegram_account)
                    
                    async def verify_password():
                        await manager.connect()
                        result = await manager.check_password(password)
                        if result['success']:
                            # Get user info
                            me = await manager.get_me()
                            await manager.disconnect()
                            return {'success': True, 'user': me}
                        await manager.disconnect()
                        return result
                    
                    result = run_async(verify_password())
                    
                    if result['success']:
                        telegram_account.is_active = True
                        telegram_account.save()
                        
                        # Clean session
                        if 'telegram_account_id' in request.session:
                            del request.session['telegram_account_id']
                        if 'phone_code_hash' in request.session:
                            del request.session['phone_code_hash']
                        
                        messages.success(request, f'Telegram account {telegram_account.phone_number} muvaffaqiyatli tasdiqlandi!')
                        return render(request, 'home/verify_telegram_account.html', {
                            'step': 'success',
                            'telegram_account': telegram_account
                        })
                    else:
                        messages.error(request, f'Noto\'g\'ri parol: {result["error"]}')
                except Exception as e:
                    messages.error(request, f'Xatolik: {str(e)}')
    
    # Determine the step for template
    step = 'send_code' if mode == 'verify' else 'reactivate'
    
    return render(request, 'home/verify_telegram_account.html', {
        'telegram_account': telegram_account,
        'step': step,
        'mode': mode,
        'is_reactivation': bool(account_id)
    })

@login_required
def contacts_list(request):
    """List all contacts"""
    account_id = request.GET.get('account')
    search = request.GET.get('search', '')
    
    contacts = Contact.objects.filter(telegram_account__user=request.user)
    
    if account_id:
        contacts = contacts.filter(telegram_account_id=account_id)
    
    if search:
        contacts = contacts.filter(
            Q(name__icontains=search) | 
            Q(phone_number__icontains=search) | 
            Q(username__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search)
        )
    
    paginator = Paginator(contacts.order_by('-created_at'), 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    telegram_accounts = TelegramAccount.objects.filter(user=request.user, is_active=True)
    
    context = {
        'page_obj': page_obj,
        'telegram_accounts': telegram_accounts,
        'current_account': account_id,
        'search': search,
    }
    return render(request, 'home/contacts_list.html', context)

@login_required
def import_contacts(request):
    """Import contacts from Excel"""
    if request.method == 'POST':
        account_id = request.POST.get('telegram_account')
        excel_file = request.FILES.get('excel_file')
        
        if not account_id:
            messages.error(request, 'Telegram accountni tanlang.')
            return redirect('import_contacts')
        
        if not excel_file:
            messages.error(request, 'Excel faylni yuklang.')
            return redirect('import_contacts')
        
        telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
        
        # Save uploaded file temporarily with unique name
        import uuid
        unique_filename = f'temp_{uuid.uuid4()}_{excel_file.name}'
        temp_file_path = os.path.join(settings.MEDIA_ROOT, unique_filename)
        os.makedirs(os.path.dirname(temp_file_path), exist_ok=True)
        
        try:
            # Write file in chunks
            with open(temp_file_path, 'wb') as temp_file:
                for chunk in excel_file.chunks():
                    temp_file.write(chunk)
            
            # Import contacts from Excel
            result = import_contacts_from_excel(temp_file_path, telegram_account)
            
            if result['success']:
                messages.success(request, 
                    f'Kontaktlar muvaffaqiyatli import qilindi! '
                    f'Jami: {result["total"]}, '
                    f'Muvaffaqiyatli: {result["successful"]}, '
                    f'Xato: {result["failed"]}'
                )
            else:
                messages.error(request, f'Import xatolik: {result["error"]}')
        
        except Exception as e:
            messages.error(request, f'Xatolik: {str(e)}')
        finally:
            # Clean up temp file with retry mechanism for Windows
            import time
            for attempt in range(3):
                try:
                    if os.path.exists(temp_file_path):
                        time.sleep(0.1)  # Wait a bit for file handles to close
                        os.remove(temp_file_path)
                    break
                except PermissionError:
                    if attempt == 2:  # Last attempt
                        print(f"Warning: Could not delete temp file {temp_file_path}")
                    time.sleep(0.5)  # Wait longer before retry
        
        return redirect('contacts_list')
    
    telegram_accounts = TelegramAccount.objects.filter(user=request.user, is_active=True)
    return render(request, 'home/import_contacts.html', {
        'telegram_accounts': telegram_accounts
    })

@login_required
def chats_list(request):
    """List chats for selected Telegram account - Telegram Style"""
    telegram_accounts = TelegramAccount.objects.filter(user=request.user, is_active=True)
    
    account_id = request.GET.get('account')
    selected_account = None
    
    # Agar account tanlanmagan bo'lsa va accountlar mavjud bo'lsa
    if not account_id and telegram_accounts.exists():
        selected_account = telegram_accounts.first()
        account_id = selected_account.id
    elif account_id:
        selected_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    chats = []
    telegram_account = selected_account
    
    # Load fresh chats from Telegram
    if request.GET.get('refresh') == '1':
        try:
            # Build coroutine that fetches dialogs and messages (no DB access inside)
            async def _fetch_dialogs_with_messages():
                manager = TelegramManager(telegram_account)
                await manager.connect()
                try:
                    if not await manager.is_authorized():
                        return {'success': False, 'error': 'Not authorized'}
                    
                    # Get dialogs
                    client = await manager.get_client()
                    dialogs = await client.get_dialogs()
                    
                    chats_data = []
                    for dialog in dialogs:
                        # Get last 20 messages for each dialog (reverse=False for newest first)
                        messages = await client.get_messages(dialog.id, limit=20)
                        
                        messages_data = []
                        for msg in messages:
                            if msg.text:  # Only text messages
                                messages_data.append({
                                    'id': msg.id,
                                    'text': msg.text,
                                    'sender_id': msg.sender_id,
                                    'date': msg.date,
                                    'is_outgoing': msg.out
                                })
                        
                        # Avatar yuklab olish
                        avatar_path = None
                        try:
                            avatar_path = await manager.download_chat_avatar(dialog.id)
                        except Exception as e:
                            print(f"Error downloading avatar for {dialog.id}: {e}")
                        
                        chats_data.append({
                            'id': dialog.id,
                            'title': dialog.name,
                            'username': getattr(dialog.entity, 'username', None),
                            'is_user': dialog.is_user,
                            'is_group': dialog.is_group,
                            'is_channel': dialog.is_channel,
                            'avatar': avatar_path,
                            'messages': messages_data
                        })
                    
                    return {'success': True, 'chats': chats_data}
                finally:
                    try:
                        await manager.disconnect()
                    except Exception:
                        pass

            # Run coroutine in a separate thread
            result = run_coroutine_in_thread(_fetch_dialogs_with_messages(), timeout=90)

            if not result.get('success'):
                messages.error(request, f"Chatlarni yuklashda xatolik: {result.get('error')}")
            else:
                coro_res = result.get('result')
                chats_data = []
                if isinstance(coro_res, dict) and coro_res.get('success'):
                    chats_data = coro_res.get('chats', [])

                # Save to database using SINGLE atomic transaction (PostgreSQL optimized)
                from django.db import transaction, connection
                
                # Close old connections
                connection.close()
                
                try:
                    # SINGLE atomic transaction for ALL operations
                    with transaction.atomic():
                        # Bulk collect existing chat IDs to avoid repeated queries
                        existing_chat_ids = set(
                            Chat.objects.filter(
                                telegram_account=telegram_account
                            ).values_list('chat_id', flat=True)
                        )
                        
                        # Prepare bulk operations
                        chats_to_create = []
                        chats_to_update = []
                        
                        for chat_data in chats_data:
                            chat_id = chat_data['id']
                            
                            if chat_id not in existing_chat_ids:
                                # New chat - prepare for bulk create
                                chats_to_create.append(Chat(
                                    telegram_account=telegram_account,
                                    chat_id=chat_id,
                                    title=chat_data.get('title', ''),
                                    username=chat_data.get('username', '') or '',
                                    chat_type='private' if chat_data.get('is_user') else 
                                             'group' if chat_data.get('is_group') else 'channel',
                                    member_count=0,
                                    avatar=chat_data.get('avatar', None),
                                ))
                            else:
                                # Existing chat - will update
                                chats_to_update.append(chat_data)
                        
                        # Bulk create new chats
                        if chats_to_create:
                            Chat.objects.bulk_create(chats_to_create, ignore_conflicts=True)
                        
                        # Update existing chats
                        for chat_data in chats_to_update:
                            Chat.objects.filter(
                                telegram_account=telegram_account,
                                chat_id=chat_data['id']
                            ).update(
                                title=chat_data.get('title', ''),
                                username=chat_data.get('username', '') or '',
                                avatar=chat_data.get('avatar', None),
                            )
                        
                        # Now save messages in bulk
                        all_messages_to_create = []
                        
                        # Get all chat objects for message linking
                        chat_objects = {
                            c.chat_id: c for c in Chat.objects.filter(
                                telegram_account=telegram_account,
                                chat_id__in=[cd['id'] for cd in chats_data]
                            )
                        }
                        
                        # Get existing message IDs to avoid duplicates
                        existing_message_ids = set(
                            Message.objects.filter(
                                telegram_account=telegram_account
                            ).values_list('message_id', flat=True)
                        )
                        
                        for chat_data in chats_data:
                            chat_obj = chat_objects.get(chat_data['id'])
                            if chat_obj:
                                for msg_data in chat_data.get('messages', []):
                                    msg_id = msg_data['id']
                                    if msg_id not in existing_message_ids:
                                        all_messages_to_create.append(Message(
                                            message_id=msg_id,
                                            chat=chat_obj,
                                            telegram_account=telegram_account,
                                            text=msg_data.get('text', ''),
                                            message_type='text',
                                            is_outgoing=msg_data.get('is_outgoing', False),
                                            date=msg_data.get('date'),
                                            sender_id=msg_data.get('sender_id'),
                                        ))
                        
                        # Bulk create all messages
                        if all_messages_to_create:
                            Message.objects.bulk_create(all_messages_to_create, batch_size=500, ignore_conflicts=True)
                    
                    messages.success(request, 'Chatlar yangilandi!')
                except Exception as e:
                    messages.error(request, f'Chatlarni yuklashda xatolik: {str(e)}')
        except Exception as e:
            messages.error(request, f'Xatolik: {str(e)}')
    
    # Get chats with their last messages (ordered by date ascending for display)
    chats = Chat.objects.filter(telegram_account=telegram_account).prefetch_related(
        Prefetch('messages', queryset=Message.objects.order_by('date'))
    ).order_by('-updated_at')
    
    # No pagination for telegram style
    telegram_accounts = TelegramAccount.objects.filter(user=request.user, is_active=True)
    
    context = {
        'chats': chats,
        'telegram_account': telegram_account,
        'telegram_accounts': telegram_accounts,
    }
    return render(request, 'home/chats_list.html', context)

@login_required
def chat_messages(request, chat_id):
    """View messages in a specific chat"""
    chat = get_object_or_404(Chat, id=chat_id, telegram_account__user=request.user)
    
    # Load messages from Telegram if requested
    if request.GET.get('load') == '1':
        try:
            async def load_messages_async():
                try:
                    manager = TelegramManager(chat.telegram_account)
                    await manager.connect()
                    try:
                        if await manager.is_authorized():
                            messages_data = await manager.get_messages(chat.chat_id, limit=100)
                            return {'success': True, 'messages': messages_data}
                        return {'success': False, 'error': 'Not authorized'}
                    finally:
                        await manager.disconnect()
                except Exception as e:
                    return {'success': False, 'error': str(e)}
            
            # Run with proper event loop management
            result_wrapper = run_coroutine_in_thread(load_messages_async(), timeout=30)
            
            if result_wrapper.get('success'):
                result = result_wrapper.get('result', {})
                
                if result.get('success'):
                    messages_data = result.get('messages', [])
                    
                    # Bulk save to database with atomic transaction
                    with transaction.atomic():
                        # Get existing message IDs to avoid duplicates
                        existing_message_ids = set(
                            Message.objects.filter(
                                chat=chat,
                                telegram_account=chat.telegram_account
                            ).values_list('message_id', flat=True)
                        )
                        
                        # Prepare new messages for bulk insert
                        new_messages = []
                        for msg_data in messages_data:
                            if msg_data['id'] not in existing_message_ids:
                                new_messages.append(Message(
                                    message_id=msg_data['id'],
                                    chat=chat,
                                    telegram_account=chat.telegram_account,
                                    text=msg_data.get('text', ''),
                                    message_type='text',
                                    is_outgoing=msg_data.get('is_outgoing', False),
                                    date=msg_data.get('date'),
                                    reply_to_message_id=msg_data.get('reply_to'),
                                ))
                        
                        # Bulk insert new messages
                        if new_messages:
                            Message.objects.bulk_create(new_messages, batch_size=100, ignore_conflicts=True)
                    
                    messages.success(request, 'Xabarlar yuklandi!')
                else:
                    messages.error(request, f'Xabarlarni yuklashda xatolik: {result.get("error")}')
            else:
                messages.error(request, f'Xabarlarni yuklashda xatolik: {result_wrapper.get("error")}')
        except Exception as e:
            messages.error(request, f'Xabarlarni yuklashda xatolik: {str(e)}')
    
    # Send message if POST request
    if request.method == 'POST':
        message_text = request.POST.get('message_text')
        if message_text:
            try:
                manager = TelegramManager(chat.telegram_account)
                
                async def send_msg():
                    await manager.connect()
                    if await manager.is_authorized():
                        result = await manager.send_message(chat.chat_id, message_text)
                        await manager.disconnect()
                        return result
                    await manager.disconnect()
                    return {'success': False, 'error': 'Not authorized'}
                
                result = run_async(send_msg())
                if result['success']:
                    messages.success(request, 'Xabar yuborildi!')
                    # Create message record
                    Message.objects.create(
                        message_id=result['message_id'],
                        chat=chat,
                        telegram_account=chat.telegram_account,
                        text=message_text,
                        message_type='text',
                        is_outgoing=True,
                        date=datetime.now(),
                    )
                else:
                    messages.error(request, f'Xabar yuborishda xatolik: {result["error"]}')
            except Exception as e:
                messages.error(request, f'Xatolik: {str(e)}')
        
        return redirect('chat_messages', chat_id=chat_id)
    
    chat_messages = Message.objects.filter(chat=chat).order_by('-date')
    paginator = Paginator(chat_messages, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'chat': chat,
        'page_obj': page_obj,
    }
    return render(request, 'home/chat_messages.html', context)

@login_required
def ai_integration(request):
    """AI Integration management"""
    ai_integrations = AIIntegration.objects.filter(telegram_account__user=request.user)
    telegram_accounts = TelegramAccount.objects.filter(user=request.user, is_active=True)
    
    if request.method == 'POST':
        # Add new AI integration
        account_id = request.POST.get('telegram_account')
        provider = request.POST.get('provider')
        model_name = request.POST.get('model_name')
        api_key = request.POST.get('api_key')
        system_prompt = request.POST.get('system_prompt')
        
        if account_id and provider and model_name:
            telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
            
            AIIntegration.objects.create(
                telegram_account=telegram_account,
                provider=provider,
                model_name=model_name,
                api_key=api_key,
                system_prompt=system_prompt or "Siz foydali AI yordamchisisiz.",
                is_active=True  # Darhol faollashtirish
            )
            messages.success(request, 'AI integration muvaffaqiyatli qo\'shildi va faollashtirildi!')
            return redirect('ai_integration')
    
    context = {
        'ai_integrations': ai_integrations,
        'telegram_accounts': telegram_accounts,
        'providers': ['openai', 'claude', 'local'],
    }
    return render(request, 'home/ai_integration.html', context)

@login_required
@require_http_methods(["POST"])
def send_ai_message(request, account_id, chat_id):
    """Send AI-powered message to chat"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    # Get AI integration for this account
    try:
        ai_integration = AIIntegration.objects.get(
            telegram_account=telegram_account,
            is_active=True
        )
    except AIIntegration.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'Bu account uchun AI integration topilmadi'
        })
    
    user_message = request.POST.get('message', '').strip()
    if not user_message:
        return JsonResponse({'success': False, 'error': 'Xabar matni bo\'sh'})
    
    try:
        # Generate AI response (bu yerda OpenAI yoki boshqa AI service ni chaqirishimiz kerak)
        ai_response = generate_ai_response(ai_integration, user_message)
        
        if ai_response:
            manager = TelegramManager(telegram_account)
            
            async def send_msg():
                await manager.connect()
                if await manager.is_authorized():
                    result = await manager.send_message(chat_id, ai_response)
                    await manager.disconnect()
                    return result
                await manager.disconnect()
                return {'success': False, 'error': 'Not authorized'}
            
            result = run_async(send_msg())
            
            if result['success']:
                return JsonResponse({
                    'success': True, 
                    'message': 'AI xabar yuborildi',
                    'ai_response': ai_response
                })
            else:
                return JsonResponse({
                    'success': False, 
                    'error': result['error']
                })
        else:
            return JsonResponse({
                'success': False, 
                'error': 'AI javob yarata olmadi'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': str(e)
        })

def generate_ai_response(ai_integration, user_message):
    """Generate AI response based on integration settings"""
    try:
        if ai_integration.provider == 'openai':
            # OpenAI integration (eski API)
            import openai
            openai.api_key = ai_integration.api_key
            
            response = openai.ChatCompletion.create(
                model=ai_integration.model_name or "gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": ai_integration.system_prompt},
                    {"role": "user", "content": user_message}
                ],
                max_tokens=500
            )
            return response.choices[0].message.content
            
        elif ai_integration.provider == 'claude':
            # Claude integration (Anthropic)
            return f"Claude javob: {user_message} haqida AI javob"
            
        elif ai_integration.provider == 'local':
            # Local AI model
            return f"Lokal AI javob: {user_message} haqida javob"
            
        else:
            return "Noma'lum AI provider"
            
    except Exception as e:
        print(f"AI response generation error: {e}")
        return None
    
    context = {
        'ai_integrations': ai_integrations,
        'telegram_accounts': telegram_accounts,
    }
    return render(request, 'home/ai_integration.html', context)

@login_required
def check_account_status(request, account_id):
    """Check if Telegram account session is still valid"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    try:
        manager = TelegramManager(telegram_account)
        
        async def check_status():
            try:
                client = await manager.get_client()
                if client.is_connected():
                    # Try to get self information to verify session
                    me = await client.get_me()
                    if me:
                        # Update account status if needed
                        if not telegram_account.is_active:
                            telegram_account.is_active = True
                            telegram_account.status = 'active'
                            telegram_account.save()
                        return {'valid': True, 'username': me.username or me.phone}
                    else:
                        return {'valid': False, 'error': 'Session expired'}
                else:
                    return {'valid': False, 'error': 'Not connected'}
            except Exception as e:
                return {'valid': False, 'error': str(e)}
        
        result = run_async(check_status())
        
        if result['valid']:
            return JsonResponse({
                'success': True,
                'status': 'active',
                'username': result.get('username', ''),
                'message': 'Account is active and connected'
            })
        else:
            # Update account status
            telegram_account.is_active = False
            telegram_account.status = 'session_expired'
            telegram_account.save()
            
            return JsonResponse({
                'success': False,
                'status': 'inactive',
                'message': f'Account session invalid: {result["error"]}'
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'status': 'error',
            'message': f'Error checking status: {str(e)}'
        })

@login_required
@require_http_methods(["POST"])
def sync_telegram_contacts(request, account_id):
    """Sync contacts from Telegram"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    try:
        manager = TelegramManager(telegram_account)
        
        async def sync_contacts():
            await manager.connect()
            if await manager.is_authorized():
                result = await manager.sync_contacts_from_telegram()
                await manager.disconnect()
                return result
            await manager.disconnect()
            return {'success': False, 'error': 'Not authorized'}
        
        result = run_async(sync_contacts())
        
        if result['success']:
            return JsonResponse({
                'success': True, 
                'message': f'{result["synced"]} ta yangi kontakt sinxronlashtirildi!'
            })
        else:
            return JsonResponse({
                'success': False, 
                'message': f'Xatolik: {result["error"]}'
            })
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'Xatolik: {str(e)}'
        })

@login_required
@require_http_methods(["POST"])
def add_contacts_to_telegram(request, account_id):
    """Add contacts from database to Telegram"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    try:
        manager = TelegramManager(telegram_account)
        
        async def add_contacts():
            await manager.connect()
            if await manager.is_authorized():
                result = await manager.add_contacts_to_telegram()
                await manager.disconnect()
                return result
            await manager.disconnect()
            return {'success': False, 'error': 'Not authorized'}
        
        result = run_async(add_contacts())
        
        if result['success']:
            message = result['message']
            if 'debug' in result:
                message += f'\n\nDebug ma\'lumoti: {result["debug"]}'
            
            return JsonResponse({
                'success': True, 
                'message': message,
                'added': result.get('added', 0),
                'imported': result.get('imported', 0),
                'debug': result.get('debug', '')
            })
        else:
            return JsonResponse({
                'success': False, 
                'message': result['error']
            })
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'Xatolik: {str(e)}'
        })

@login_required
@require_http_methods(["POST"])
def fix_phone_numbers(request, account_id):
    """Fix phone numbers format in database"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    try:
        # Get contacts with potentially invalid phone numbers
        contacts = Contact.objects.filter(
            telegram_account=telegram_account
        ).exclude(
            phone_number__isnull=True
        ).exclude(
            phone_number__exact=''
        )
        
        fixed_count = 0
        for contact in contacts:
            original_phone = contact.phone_number
            digits_only = ''.join(filter(str.isdigit, original_phone))
            
            new_phone = None
            if len(digits_only) == 9 and digits_only.startswith('9'):
                # Uzbekistan mobile: 9XXXXXXXX -> +9989XXXXXXXX
                new_phone = '+998' + digits_only
            elif len(digits_only) == 10 and not original_phone.startswith('+'):
                # Assume Russia/Kazakhstan: XXXXXXXXXX -> +7XXXXXXXXXX  
                new_phone = '+7' + digits_only
            elif len(digits_only) == 12 and digits_only.startswith('998'):
                # Already correct Uzbekistan format
                new_phone = '+' + digits_only
            elif len(digits_only) == 11 and digits_only.startswith('7'):
                # Already correct Russia format
                new_phone = '+' + digits_only
            elif len(digits_only) >= 9 and not original_phone.startswith('+'):
                # Generic case - add + sign
                # For 9-digit numbers starting with 9, assume Uzbekistan
                if len(digits_only) == 9 and digits_only.startswith('9'):
                    new_phone = '+998' + digits_only
                else:
                    new_phone = '+' + digits_only
            
            if new_phone and new_phone != original_phone:
                contact.phone_number = new_phone
                contact.save()
                fixed_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'{fixed_count} ta telefon raqami to\'g\'irlandi',
            'fixed': fixed_count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Xatolik: {str(e)}'
        })

@login_required
@require_http_methods(["POST"])
def reset_contacts_status(request, account_id):
    """Reset all contacts to not added status"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    try:
        # Reset all contacts to not added
        updated_count = Contact.objects.filter(
            telegram_account=telegram_account
        ).update(added_to_telegram=False)
        
        return JsonResponse({
            'success': True,
            'message': f'{updated_count} ta kontakt holati reset qilindi',
            'reset': updated_count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Xatolik: {str(e)}'
        })

@login_required
def chats_detail(request, account_id, chat_id):
    """Chat details and messages"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    # Get or create chat record
    chat, created = Chat.objects.get_or_create(
        telegram_account=telegram_account,
        chat_id=chat_id,
        defaults={'title': f'Chat {chat_id}', 'chat_type': 'private'}
    )
    
    # Get messages from database (ordered oldest to newest for display)
    db_messages = Message.objects.filter(
        chat=chat,
        telegram_account=telegram_account
    ).order_by('date')
    
    # Load fresh messages from Telegram if requested
    if request.GET.get('load') == '1' or not db_messages.exists():
        try:
            async def _fetch_messages():
                manager = TelegramManager(telegram_account)
                await manager.connect()
                try:
                    if await manager.is_authorized():
                        client = await manager.get_client()
                        messages = await client.get_messages(chat_id, limit=100)
                        
                        messages_data = []
                        for msg in messages:
                            if msg.text:
                                messages_data.append({
                                    'id': msg.id,
                                    'text': msg.text,
                                    'sender_id': msg.sender_id,
                                    'date': msg.date,
                                    'is_outgoing': msg.out
                                })
                        return {'success': True, 'messages': messages_data}
                    return {'success': False, 'error': 'Not authorized'}
                finally:
                    try:
                        await manager.disconnect()
                    except Exception:
                        pass
            
            result = run_coroutine_in_thread(_fetch_messages(), timeout=60)
            
            if result.get('success'):
                coro_res = result.get('result')
                if isinstance(coro_res, dict) and coro_res.get('success'):
                    messages_data = coro_res.get('messages', [])
                    
                    # Bulk save to database with atomic transaction
                    with transaction.atomic():
                        # Get existing message IDs to avoid duplicates
                        existing_message_ids = set(
                            Message.objects.filter(
                                chat=chat,
                                telegram_account=telegram_account
                            ).values_list('message_id', flat=True)
                        )
                        
                        # Prepare new messages for bulk insert
                        new_messages = []
                        for msg_data in messages_data:
                            if msg_data['id'] not in existing_message_ids:
                                new_messages.append(Message(
                                    message_id=msg_data['id'],
                                    chat=chat,
                                    telegram_account=telegram_account,
                                    text=msg_data.get('text', ''),
                                    message_type='text',
                                    is_outgoing=msg_data.get('is_outgoing', False),
                                    date=msg_data.get('date'),
                                    sender_id=msg_data.get('sender_id'),
                                ))
                        
                        # Bulk insert new messages
                        if new_messages:
                            Message.objects.bulk_create(new_messages, batch_size=100, ignore_conflicts=True)
                    
                    messages.success(request, 'Xabarlar yangilandi!')
                    # Reload messages from database
                    db_messages = Message.objects.filter(
                        chat=chat,
                        telegram_account=telegram_account
                    ).order_by('date')
            else:
                messages.error(request, f'Xatolik: {result.get("error")}')
        except Exception as e:
            messages.error(request, f'Xatolik: {str(e)}')
    
    context = {
        'telegram_account': telegram_account,
        'chat': chat,
        'messages': db_messages,
    }
    return render(request, 'home/chat_detail.html', context)

@login_required
@require_http_methods(["POST"])
def send_message_to_chat(request, account_id, chat_id):
    """Send message to a chat"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    # Chat obyektini olish - chat_id bu database ID, telegram_chat_id kerak
    chat = get_object_or_404(Chat, id=chat_id, telegram_account=telegram_account)
    telegram_chat_id = chat.chat_id  # Bu real Telegram chat ID
    
    message_text = request.POST.get('message', '').strip()
    if not message_text:
        return JsonResponse({'success': False, 'error': 'Xabar matni bo\'sh'})
    
    try:
        # Use run_coroutine_in_thread helper
        async def send_msg_async():
            try:
                manager = TelegramManager(telegram_account)
                await manager.connect()
                try:
                    if await manager.is_authorized():
                        result = await manager.send_message(telegram_chat_id, message_text)
                        return result
                    return {'success': False, 'error': 'Not authorized'}
                finally:
                    await manager.disconnect()
            except Exception as e:
                return {'success': False, 'error': str(e)}
        
        # Run with proper event loop management
        result_wrapper = run_coroutine_in_thread(send_msg_async(), timeout=30)
        
        if not result_wrapper.get('success'):
            return JsonResponse({'success': False, 'error': result_wrapper.get('error', 'Unknown error')})
        
        result = result_wrapper.get('result', {})
        
        if result.get('success'):
            # Save message to database with atomic transaction
            from django.utils import timezone
            with transaction.atomic():
                Message.objects.create(
                    telegram_account=telegram_account,
                    chat=chat,
                    message_id=result['message_id'],
                    text=message_text,
                    is_outgoing=True,
                    sender_id=telegram_account.user.id,
                    date=timezone.now(),
                    message_type='text'
                )
            
            return JsonResponse({
                'success': True, 
                'message': 'Xabar yuborildi',
                'message_id': result['message_id']
            })
        else:
            return JsonResponse({
                'success': False, 
                'error': result['error']
            })
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': str(e)
        })

@login_required
@require_http_methods(["POST"])
def send_ai_message_to_chat(request, account_id, chat_id):
    """Send AI-powered message to a specific chat"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    # Get AI integration for this account
    try:
        ai_integration = AIIntegration.objects.get(
            telegram_account=telegram_account,
            is_active=True
        )
    except AIIntegration.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'Bu account uchun faol AI integration topilmadi'
        })
    
    user_prompt = request.POST.get('prompt', '').strip()
    if not user_prompt:
        return JsonResponse({'success': False, 'error': 'AI prompt bo\'sh'})
    
    try:
        # Generate AI response
        ai_response = generate_advanced_ai_response(ai_integration, user_prompt)
        
        if ai_response:
            async def send_ai_msg_async():
                try:
                    manager = TelegramManager(telegram_account)
                    await manager.connect()
                    try:
                        if await manager.is_authorized():
                            result = await manager.send_message(chat_id, ai_response)
                            return result
                        return {'success': False, 'error': 'Not authorized'}
                    finally:
                        await manager.disconnect()
                except Exception as e:
                    return {'success': False, 'error': str(e)}
            
            # Run with proper event loop management
            result_wrapper = run_coroutine_in_thread(send_ai_msg_async(), timeout=30)
            
            if result_wrapper.get('success'):
                result = result_wrapper.get('result', {})
                
                if result.get('success'):
                    # Get or create chat object and save message with atomic transaction
                    with transaction.atomic():
                        chat, created = Chat.objects.get_or_create(
                            telegram_account=telegram_account,
                            chat_id=chat_id,
                            defaults={'title': f'Chat {chat_id}', 'chat_type': 'private'}
                        )
                        
                        # Save AI message to database
                        Message.objects.create(
                            telegram_account=telegram_account,
                            chat=chat,
                            message_id=result['message_id'],
                            text=ai_response,
                            is_outgoing=True,
                            sender_id=telegram_account.user.id,
                            date=timezone.now(),
                            message_type='text'
                        )
                
                    return JsonResponse({
                        'success': True, 
                        'message': 'AI xabar muvaffaqiyatli yuborildi',
                        'ai_response': ai_response,
                        'message_id': result['message_id']
                    })
                else:
                    return JsonResponse({
                        'success': False, 
                        'error': f'Telegram xabar yuborishda xatolik: {result.get("error")}'
                    })
            else:
                return JsonResponse({
                    'success': False, 
                    'error': f'Xatolik: {result_wrapper.get("error")}'
                })
        else:
            return JsonResponse({
                'success': False, 
                'error': 'AI javob yarata olmadi. AI sozlamalarni tekshiring.'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'AI xabar yuborishda xatolik: {str(e)}'
        })

def generate_advanced_ai_response(ai_integration, user_prompt):
    """Advanced AI response generator with multiple providers"""
    try:
        if ai_integration.provider == 'openai':
            # OpenAI GPT integration (eski API)
            try:
                import openai
                openai.api_key = ai_integration.api_key
                
                response = openai.ChatCompletion.create(
                    model=ai_integration.model_name or "gpt-3.5-turbo",
                    messages=[
                        {"role": "system", "content": ai_integration.system_prompt or "Siz foydali AI yordamchisisiz."},
                        {"role": "user", "content": user_prompt}
                    ],
                    max_tokens=1000,
                    temperature=0.7
                )
                return response.choices[0].message.content
            except Exception as e:
                print(f"OpenAI API error: {e}")
                return f"OpenAI xatolik: {str(e)}"
                
        elif ai_integration.provider == 'claude':
            # Claude/Anthropic integration (placeholder)
            try:
                # Bu yerda Claude API ni chaqirish kerak
                # import anthropic
                # client = anthropic.Anthropic(api_key=ai_integration.api_key)
                # response = client.messages.create(...)
                return f"Claude AI javobi: {user_prompt} haqida professional javob. (Demo rejim)"
            except Exception as e:
                return f"Claude xatolik: {str(e)}"
            
        elif ai_integration.provider == 'local':
            # Local AI model integration
            try:
                # Bu yerda lokal AI model (masalan, Ollama) ni chaqirish kerak
                return f"Lokal AI javobi: {user_prompt} haqida lokal model javobi. (Demo rejim)"
            except Exception as e:
                return f"Lokal AI xatolik: {str(e)}"
            
        else:
            return f"Noma'lum AI provider: {ai_integration.provider}"
            
    except Exception as e:
        print(f"AI response generation error: {e}")
        return None

@login_required
def ai_chat_interface(request, account_id):
    """AI chat interface for testing AI responses"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    # Get AI integration
    try:
        ai_integration = AIIntegration.objects.get(
            telegram_account=telegram_account,
            is_active=True
        )
    except AIIntegration.DoesNotExist:
        messages.error(request, 'Bu account uchun AI integration mavjud emas')
        return redirect('ai_integration')
    
    ai_history = []
    
    if request.method == 'POST':
        user_input = request.POST.get('user_input', '').strip()
        
        if user_input:
            # Generate AI response
            ai_response = generate_advanced_ai_response(ai_integration, user_input)
            
            ai_history = [
                {'type': 'user', 'message': user_input},
                {'type': 'ai', 'message': ai_response or 'AI javob yaratilmadi'}
            ]
    
    context = {
        'telegram_account': telegram_account,
        'ai_integration': ai_integration,
        'ai_history': ai_history,
    }
    return render(request, 'home/ai_chat.html', context)

@login_required
@require_http_methods(["POST"])
def test_ai_integration(request, account_id):
    """Test AI integration without sending to Telegram"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    # Get AI integration
    try:
        ai_integration = AIIntegration.objects.get(
            telegram_account=telegram_account,
            is_active=True
        )
    except AIIntegration.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'AI integration topilmadi'
        })
    
    test_prompt = request.POST.get('prompt', 'Salom, bu test xabari.')
    
    try:
        # Test AI response
        ai_response = generate_advanced_ai_response(ai_integration, test_prompt)
        
        if ai_response:
            return JsonResponse({
                'success': True, 
                'prompt': test_prompt,
                'response': ai_response,
                'provider': ai_integration.provider,
                'model': ai_integration.model_name
            })
        else:
            return JsonResponse({
                'success': False, 
                'error': 'AI javob yaratilmadi'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Test xatolik: {str(e)}'
        })

@login_required
def ai_provider_setup(request, provider):
    """Setup specific AI provider"""
    valid_providers = {
        'openai': {
            'name': 'OpenAI (GPT)',
            'icon': 'fas fa-robot',
            'color': 'success',
            'models': ['gpt-3.5-turbo', 'gpt-4', 'gpt-4-turbo'],
            'api_docs': 'https://platform.openai.com/api-keys',
            'description': 'OpenAI GPT modellaridan foydalaning. API kalitni olish uchun OpenAI platformasiga ro\'yxatdan o\'ting.'
        },
        'claude': {
            'name': 'Claude (Anthropic)',
            'icon': 'fas fa-microchip',
            'color': 'info',
            'models': ['claude-3-sonnet', 'claude-3-opus', 'claude-3-haiku'],
            'api_docs': 'https://console.anthropic.com/settings/keys',
            'description': 'Anthropic Claude AI yordamchisi. Professional va ishonchli AI javoblar uchun.'
        },
        'gemini': {
            'name': 'Gemini (Google)',
            'icon': 'fab fa-google',
            'color': 'danger',
            'models': ['gemini-pro', 'gemini-pro-vision', 'gemini-1.5-pro'],
            'api_docs': 'https://makersuite.google.com/app/apikey',
            'description': 'Google Gemini AI modeli. Ko\'p tillarda ishlaydi va vizual ma\'lumotlarni tushunadi.'
        }
    }
    
    if provider not in valid_providers:
        messages.error(request, 'Noma\'lum AI provider')
        return redirect('ai_integration')
    
    provider_info = valid_providers[provider]
    telegram_accounts = TelegramAccount.objects.filter(user=request.user, is_active=True)
    
    # Get existing integration for this provider
    existing_integration = AIIntegration.objects.filter(
        telegram_account__user=request.user,
        provider=provider
    ).first()
    
    if request.method == 'POST':
        account_id = request.POST.get('telegram_account')
        model_name = request.POST.get('model_name')
        api_key = request.POST.get('api_key')
        system_prompt = request.POST.get('system_prompt')
        
        if not all([account_id, model_name, api_key]):
            messages.error(request, 'Barcha maydonlar to\'ldirilishi kerak.')
        else:
            telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
            
            # Test API key before saving
            test_result = test_ai_provider_connection(provider, api_key, model_name)
            
            if test_result['success']:
                # Save or update integration
                integration, created = AIIntegration.objects.get_or_create(
                    telegram_account=telegram_account,
                    provider=provider,
                    defaults={
                        'model_name': model_name,
                        'api_key': api_key,
                        'system_prompt': system_prompt or f"Siz {provider_info['name']} yordamchisisiz.",
                        'is_active': True
                    }
                )
                
                if not created:
                    integration.model_name = model_name
                    integration.api_key = api_key
                    integration.system_prompt = system_prompt or integration.system_prompt
                    integration.is_active = True
                    integration.save()
                
                messages.success(request, f'{provider_info["name"]} muvaffaqiyatli sozlandi!')
                return redirect('ai_integration')
            else:
                messages.error(request, f'API test muvaffaqiyatsiz: {test_result["error"]}')
    
    context = {
        'provider': provider,
        'provider_info': provider_info,
        'telegram_accounts': telegram_accounts,
        'existing_integration': existing_integration,
    }
    return render(request, 'home/ai_provider_setup.html', context)

def test_ai_provider_connection(provider, api_key, model_name):
    """Test AI provider API connection"""
    import httpx
    
    try:
        if provider == 'openai':
            # Test OpenAI API
            headers = {
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json'
            }
            data = {
                'model': model_name or 'gpt-3.5-turbo',
                'messages': [{'role': 'user', 'content': 'Hello'}],
                'max_tokens': 5
            }
            
            with httpx.Client(timeout=10.0) as client:
                response = client.post(
                    'https://api.openai.com/v1/chat/completions',
                    headers=headers,
                    json=data
                )
                
                if response.status_code == 200:
                    result = response.json()
                    message = result['choices'][0]['message']['content']
                    return {
                        'success': True, 
                        'message': f'✓ OpenAI ulanish muvaffaqiyatli! Test javob: "{message}"',
                        'response': message
                    }
                elif response.status_code == 401:
                    return {'success': False, 'error': '❌ API kalit noto\'g\'ri'}
                elif response.status_code == 429:
                    return {'success': False, 'error': '⚠️ Rate limit: Juda ko\'p so\'rov yuborildi'}
                else:
                    error_data = response.json()
                    error_msg = error_data.get('error', {}).get('message', 'Noma\'lum xatolik')
                    return {'success': False, 'error': f'❌ {error_msg}'}
                    
        elif provider == 'claude':
            # Test Claude API (Anthropic)
            headers = {
                'x-api-key': api_key,
                'anthropic-version': '2023-06-01',
                'Content-Type': 'application/json'
            }
            data = {
                'model': model_name or 'claude-sonnet-4-20250514',
                'messages': [{'role': 'user', 'content': 'Hello'}],
                'max_tokens': 10
            }
            
            with httpx.Client(timeout=10.0) as client:
                response = client.post(
                    'https://api.anthropic.com/v1/messages',
                    headers=headers,
                    json=data
                )
                
                if response.status_code == 200:
                    result = response.json()
                    message = result['content'][0]['text']
                    return {
                        'success': True, 
                        'message': f'✓ Claude ulanish muvaffaqiyatli! Test javob: "{message}"',
                        'response': message
                    }
                elif response.status_code == 401:
                    return {'success': False, 'error': '❌ API kalit noto\'g\'ri'}
                elif response.status_code == 429:
                    return {'success': False, 'error': '⚠️ Rate limit: Juda ko\'p so\'rov yuborildi'}
                else:
                    error_data = response.json()
                    error_msg = error_data.get('error', {}).get('message', 'Noma\'lum xatolik')
                    return {'success': False, 'error': f'❌ {error_msg}'}
                
        elif provider == 'gemini':
            # Test Gemini API (Google)
            with httpx.Client(timeout=10.0) as client:
                response = client.post(
                    f'https://generativelanguage.googleapis.com/v1beta/models/{model_name or "gemini-pro"}:generateContent?key={api_key}',
                    headers={'Content-Type': 'application/json'},
                    json={
                        'contents': [{
                            'parts': [{'text': 'Hello'}]
                        }]
                    }
                )
                
                if response.status_code == 200:
                    result = response.json()
                    message = result['candidates'][0]['content']['parts'][0]['text']
                    return {
                        'success': True, 
                        'message': f'✓ Gemini ulanish muvaffaqiyatli! Test javob: "{message}"',
                        'response': message
                    }
                elif response.status_code == 400:
                    error_data = response.json()
                    error_msg = error_data.get('error', {}).get('message', 'API kalit noto\'g\'ri')
                    return {'success': False, 'error': f'❌ {error_msg}'}
                else:
                    return {'success': False, 'error': f'❌ HTTP {response.status_code}: {response.text[:100]}'}
                
        else:
            return {'success': False, 'error': '❌ Noma\'lum provider'}
            
    except httpx.ConnectTimeout:
        return {'success': False, 'error': '⏱️ Ulanish vaqti tugadi (timeout)'}
    except httpx.ReadTimeout:
        return {'success': False, 'error': '⏱️ Javob kutish vaqti tugadi'}
    except Exception as e:
        return {'success': False, 'error': f'❌ Test xatolik: {str(e)}'}

@login_required
@require_http_methods(["POST"])
def test_ai_provider(request, provider):
    """Test AI provider with AJAX"""
    api_key = request.POST.get('api_key')
    model_name = request.POST.get('model_name')
    
    if not api_key:
        return JsonResponse({'success': False, 'error': 'API kalit kiritilmagan'})
    
    result = test_ai_provider_connection(provider, api_key, model_name)
    return JsonResponse(result)

@login_required
def telegram_web_client(request):
    """Telegram web client interface"""
    # Get account parameter
    account_id = request.GET.get('account')
    if not account_id:
        # Show account selection
        telegram_accounts = TelegramAccount.objects.filter(user=request.user, is_active=True)
        return render(request, 'home/telegram_web_client.html', {
            'telegram_accounts': telegram_accounts,
            'show_account_selection': True
        })
    
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    # Get chats for this account
    chats = Chat.objects.filter(telegram_account=telegram_account).order_by('-updated_at')
    
    # Load fresh chats from Telegram if requested
    if request.GET.get('refresh') == '1':
        try:
            async def load_chats_async():
                try:
                    manager = TelegramManager(telegram_account)
                    await manager.connect()
                    try:
                        if await manager.is_authorized():
                            dialogs = await manager.get_dialogs()
                            chat_data = []
                            for dialog in dialogs:
                                chat_data.append({
                                    'chat_id': dialog['id'],
                                    'title': dialog['title'],
                                    'username': dialog.get('username', ''),
                                    'chat_type': 'private' if dialog['is_user'] else 
                                               'group' if dialog['is_group'] else 'channel',
                                    'member_count': 0,
                                })
                            return {'success': True, 'chats': chat_data}
                        return {'success': False, 'error': 'Not authorized'}
                    finally:
                        await manager.disconnect()
                except Exception as e:
                    return {'success': False, 'error': str(e)}
            
            # Run with proper event loop management
            result_wrapper = run_coroutine_in_thread(load_chats_async(), timeout=30)
            
            if result_wrapper.get('success'):
                result = result_wrapper.get('result', {})
                
                if result.get('success'):
                    chats_data = result.get('chats', [])
                    
                    # Bulk save/update chats with atomic transaction
                    with transaction.atomic():
                        # Get existing chat IDs
                        existing_chats = {
                            chat.chat_id: chat 
                            for chat in Chat.objects.filter(
                                telegram_account=telegram_account
                            ).select_for_update()
                        }
                        
                        # Separate new chats and existing chats
                        new_chats = []
                        chats_to_update = []
                        
                        for chat_data in chats_data:
                            chat_id = chat_data['chat_id']
                            if chat_id in existing_chats:
                                # Update existing
                                chat_obj = existing_chats[chat_id]
                                chat_obj.title = chat_data['title']
                                chat_obj.username = chat_data['username']
                                chats_to_update.append(chat_obj)
                            else:
                                # Prepare for bulk create
                                new_chats.append(Chat(
                                    chat_id=chat_id,
                                    telegram_account=telegram_account,
                                    title=chat_data['title'],
                                    username=chat_data['username'],
                                    chat_type=chat_data['chat_type'],
                                    member_count=chat_data['member_count']
                                ))
                        
                        # Bulk create new chats
                        if new_chats:
                            Chat.objects.bulk_create(new_chats, batch_size=500, ignore_conflicts=True)
                        
                        # Bulk update existing chats
                        if chats_to_update:
                            Chat.objects.bulk_update(chats_to_update, ['title', 'username'], batch_size=500)
                    
                    messages.success(request, 'Chatlar yangilandi!')
                    chats = Chat.objects.filter(telegram_account=telegram_account).order_by('-updated_at')
                else:
                    messages.error(request, f'Chatlarni yuklashda xatolik: {result.get("error")}')
            else:
                messages.error(request, f'Chatlarni yuklashda xatolik: {result_wrapper.get("error")}')
                
        except Exception as e:
            messages.error(request, f'Chatlarni yuklashda xatolik: {str(e)}')
    
    # Pagination
    paginator = Paginator(chats, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'telegram_account': telegram_account,
        'page_obj': page_obj,
    }
    return render(request, 'home/telegram_web_client.html', context)

@login_required
def get_chat_messages_api(request, account_id, chat_id):
    """API endpoint to get chat messages"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    chat = get_object_or_404(Chat, id=chat_id, telegram_account=telegram_account)
    
    try:
        async def get_messages_async():
            try:
                manager = TelegramManager(telegram_account)
                await manager.connect()
                try:
                    if await manager.is_authorized():
                        messages_data = await manager.get_messages(chat.chat_id, limit=50)
                        return {'success': True, 'messages': messages_data}
                    return {'success': False, 'error': 'Not authorized'}
                finally:
                    await manager.disconnect()
            except Exception as e:
                return {'success': False, 'error': str(e)}
        
        # Run with proper event loop management
        result_wrapper = run_coroutine_in_thread(get_messages_async(), timeout=30)
        
        if result_wrapper.get('success'):
            result = result_wrapper.get('result', {})
            
            if result.get('success'):
                # Format messages first (no DB access)
                formatted_messages = []
                for msg_data in result['messages']:
                    formatted_messages.append({
                        'id': msg_data['id'],
                        'text': msg_data['text'],
                        'is_outgoing': msg_data['is_outgoing'],
                        'date': msg_data['date'].strftime('%Y-%m-%d %H:%M:%S'),
                        'sender_id': msg_data.get('sender_id'),
                    })
            
                # Save to database immediately (PostgreSQL is fast, no background needed)
                try:
                    # Single atomic transaction
                    with transaction.atomic():
                        # Get existing message IDs - single query
                        existing_ids = set(Message.objects.filter(
                            telegram_account=telegram_account,
                            chat=chat
                        ).values_list('message_id', flat=True))
                        
                        # Prepare new messages
                        new_messages = [
                            Message(
                                message_id=msg_data['id'],
                                telegram_account=telegram_account,
                                chat=chat,
                                text=msg_data['text'],
                                message_type='text',
                                is_outgoing=msg_data['is_outgoing'],
                                date=msg_data['date'],
                                sender_id=msg_data.get('sender_id'),
                            )
                            for msg_data in result['messages']
                            if msg_data['id'] not in existing_ids
                        ]
                        
                        # Bulk create - single query
                        if new_messages:
                            Message.objects.bulk_create(new_messages, batch_size=100, ignore_conflicts=True)
                            
                except Exception as e:
                    print(f"Error saving messages: {e}")
                
                return JsonResponse({
                    'success': True,
                    'chat': {
                        'id': chat.id,
                        'chat_id': chat.chat_id,
                        'title': chat.title or chat.username or 'Unknown',
                        'username': chat.username,
                        'chat_type': chat.chat_type,
                        'member_count': chat.member_count,
                    },
                    'messages': formatted_messages
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': result.get('error')
                })
        else:
            return JsonResponse({
                'success': False,
                'error': result_wrapper.get('error')
            })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

def get_chat_info_api(request, account_id, chat_id):
    """API endpoint to get chat information with full details from Telegram"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    chat = get_object_or_404(Chat, id=chat_id, telegram_account=telegram_account)
    
    # Telegram dan to'liq ma'lumotlarni olish
    details = {}
    try:
        async def fetch_details():
            try:
                manager = TelegramManager(telegram_account)
                await manager.connect()
                try:
                    if await manager.is_authorized():
                        entity = await manager.client.get_entity(chat.chat_id)
                        
                        details_data = {}
                        if hasattr(entity, 'premium'):
                            details_data['premium'] = entity.premium
                        if hasattr(entity, 'verified'):
                            details_data['verified'] = entity.verified
                        if hasattr(entity, 'scam'):
                            details_data['scam'] = entity.scam
                        if hasattr(entity, 'fake'):
                            details_data['fake'] = entity.fake
                        if hasattr(entity, 'bot'):
                            details_data['is_bot'] = entity.bot
                        if hasattr(entity, 'phone'):
                            details_data['phone'] = entity.phone
                        if hasattr(entity, 'lang_code'):
                            details_data['lang_code'] = entity.lang_code
                        if hasattr(entity, 'status'):
                            status = entity.status
                            if hasattr(status, '__class__'):
                                status_name = status.__class__.__name__
                                if status_name == 'UserStatusOnline':
                                    details_data['status'] = 'Online'
                                elif status_name == 'UserStatusOffline':
                                    if hasattr(status, 'was_online'):
                                        # UTC vaqtini Tashkent vaqtiga o'tkazish (+5 soat)
                                        from datetime import timedelta
                                        tashkent_time = status.was_online + timedelta(hours=5)
                                        details_data['status'] = f"Offline ({tashkent_time.strftime('%H:%M')})"
                                    else:
                                        details_data['status'] = 'Offline'
                                elif status_name == 'UserStatusRecently':
                                    details_data['status'] = 'Yaqinda'
                                elif status_name == 'UserStatusLastWeek':
                                    details_data['status'] = 'O\'tgan hafta'
                                elif status_name == 'UserStatusLastMonth':
                                    details_data['status'] = 'O\'tgan oy'
                        
                        return {'success': True, 'details': details_data}
                    return {'success': False}
                finally:
                    await manager.disconnect()
            except Exception as e:
                print(f"Error getting entity details: {e}")
                return {'success': False}
        
        # Run with proper event loop management
        result_wrapper = run_coroutine_in_thread(fetch_details(), timeout=10)
        
        if result_wrapper.get('success'):
            result = result_wrapper.get('result', {})
            if result.get('success'):
                details = result.get('details', {})
    except Exception as e:
        print(f"Error fetching entity details: {e}")
    
    return JsonResponse({
        'success': True,
        'chat': {
            'id': chat.id,
            'chat_id': chat.chat_id,
            'title': chat.title or chat.username or 'Unknown',
            'username': chat.username,
            'chat_type': chat.chat_type,
            'member_count': chat.member_count,
            'avatar': chat.avatar,
            'messages_count': chat.messages.count(),
            'updated_at': chat.updated_at.isoformat(),
            'created_at': chat.created_at.isoformat(),
        },
        'details': details
    })


@login_required
def check_new_messages(request, account_id):
    """Check for new messages (polling endpoint)"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    # Get last check timestamp from request
    last_check = request.GET.get('since')
    if last_check:
        try:
            from django.utils.dateparse import parse_datetime
            last_check_time = parse_datetime(last_check)
        except:
            last_check_time = timezone.now() - timedelta(seconds=30)
    else:
        last_check_time = timezone.now() - timedelta(seconds=30)
    
    # Get new messages since last check
    new_messages = Message.objects.filter(
        telegram_account=telegram_account,
        date__gt=last_check_time,
        is_outgoing=False  # Only incoming messages
    ).select_related('chat').order_by('-date')[:50]
    
    # Format messages
    messages_data = []
    for msg in new_messages:
        messages_data.append({
            'id': msg.id,
            'message_id': msg.message_id,
            'chat_id': msg.chat.chat_id,
            'chat_title': msg.chat.title,
            'text': msg.text,
            'sender_id': msg.sender_id,
            'date': msg.date.isoformat(),
            'is_outgoing': msg.is_outgoing,
        })
    
    return JsonResponse({
        'success': True,
        'new_messages': messages_data,
        'count': len(messages_data),
        'timestamp': timezone.now().isoformat()
    })


@login_required  
def get_unread_count(request):
    """Get unread message count for all accounts"""
    accounts = TelegramAccount.objects.filter(user=request.user, is_active=True)
    
    last_check = request.GET.get('since')
    if last_check:
        try:
            from django.utils.dateparse import parse_datetime
            last_check_time = parse_datetime(last_check)
        except:
            last_check_time = timezone.now() - timedelta(minutes=5)
    else:
        last_check_time = timezone.now() - timedelta(minutes=5)
    
    result = {}
    total_unread = 0
    
    for account in accounts:
        unread_count = Message.objects.filter(
            telegram_account=account,
            date__gt=last_check_time,
            is_outgoing=False
        ).count()
        
        result[account.id] = unread_count
        total_unread += unread_count
    
    return JsonResponse({
        'success': True,
        'accounts': result,
        'total': total_unread,
        'timestamp': timezone.now().isoformat()
    })

@login_required
def chats_list_new(request):
    """New Telegram-style chats list"""
    accounts = TelegramAccount.objects.filter(
        user=request.user,
        is_active=True
    ).order_by('-created_at')
    
    context = {
        'accounts': accounts,
    }
    return render(request, 'home/chats_list_new.html', context)

@login_required
def get_chats_api(request, account_id):
    """API to get chats for an account"""
    try:
        account = TelegramAccount.objects.get(id=account_id, user=request.user)
        
        chats = Chat.objects.filter(
            telegram_account=account
        ).order_by('-updated_at')
        
        chats_data = []
        for chat in chats:
            chats_data.append({
                'id': chat.id,
                'chat_id': chat.chat_id,
                'title': chat.title,
                'username': chat.username,
                'chat_type': chat.chat_type,
                'member_count': chat.member_count,
                'is_verified': chat.is_verified,
            })
        
        return JsonResponse({
            'success': True,
            'chats': chats_data
        })
    
    except TelegramAccount.DoesNotExist:
        return JsonResponse({'error': 'Account not found'}, status=404)
    except Exception as e:
        logger.error(f"Error getting chats: {e}")
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def telegram_account_settings(request, account_id):
    """Telegram account settings page"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    # Get Telegram profile info
    profile_info = None
    try:
        async def get_profile_info():
            try:
                manager = TelegramManager(telegram_account)
                await manager.connect()
                try:
                    if await manager.is_authorized():
                        from telethon.tl.functions.users import GetFullUserRequest
                        
                        me = await manager.client.get_me()
                        full = await manager.client(GetFullUserRequest(me.id))
                        
                        return {
                            'success': True,
                            'first_name': me.first_name or '',
                            'last_name': me.last_name or '',
                            'username': me.username or '',
                            'phone': me.phone or '',
                            'bio': full.full_user.about or '',
                            'premium': me.premium or False,
                            'verified': me.verified or False,
                            'lang_code': me.lang_code or '',
                            'photo_id': me.photo.photo_id if me.photo else None,
                        }
                    return {'success': False, 'error': 'Not authorized'}
                finally:
                    await manager.disconnect()
            except Exception as e:
                return {'success': False, 'error': str(e)}
        
        result = run_coroutine_in_thread(get_profile_info(), timeout=30)
        if result.get('success'):
            profile_info = result.get('result', {})
    except Exception as e:
        print(f"Error getting profile info: {e}")
    
    return render(request, 'home/telegram_account_settings.html', {
        'telegram_account': telegram_account,
        'profile_info': profile_info
    })

@login_required
@require_http_methods(["POST"])
def delete_all_contacts(request, account_id):
    """Delete all contacts from Telegram account"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    try:
        async def delete_contacts_async():
            try:
                manager = TelegramManager(telegram_account)
                await manager.connect()
                try:
                    if not await manager.is_authorized():
                        return {'success': False, 'error': 'Account avtorizatsiya qilinmagan'}
                    
                    # Get all contacts
                    from telethon.tl.functions.contacts import GetContactsRequest, DeleteContactsRequest
                    
                    result = await manager.client(GetContactsRequest(hash=0))
                    
                    if not hasattr(result, 'users') or not result.users:
                        return {'success': True, 'message': 'Kontaktlar topilmadi', 'deleted_count': 0}
                    
                    # Get contact IDs
                    contact_ids = [user.id for user in result.users]
                    
                    # Delete all contacts
                    await manager.client(DeleteContactsRequest(id=contact_ids))
                    
                    return {
                        'success': True, 
                        'message': f'{len(contact_ids)} ta kontakt o\'chirildi!',
                        'deleted_count': len(contact_ids)
                    }
                    
                finally:
                    await manager.disconnect()
                    
            except Exception as e:
                return {'success': False, 'error': str(e)}
        
        # Run with proper event loop management
        result_wrapper = run_coroutine_in_thread(delete_contacts_async(), timeout=60)
        
        if result_wrapper.get('success'):
            result = result_wrapper.get('result', {})
            if result.get('success'):
                return JsonResponse({
                    'success': True,
                    'message': result.get('message'),
                    'deleted_count': result.get('deleted_count', 0)
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': result.get('error', 'Noma\'lum xatolik')
                })
        else:
            return JsonResponse({
                'success': False,
                'error': result_wrapper.get('error', 'Noma\'lum xatolik')
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@require_http_methods(["POST"])
def delete_telegram_account(request, account_id):
    """Delete Telegram account completely (session, database data)"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    try:
        # Delete session file
        import os
        session_files = [
            os.path.join('telegram_sessions', f'{telegram_account.phone_number}.session'),
            os.path.join(settings.BASE_DIR, 'telegram_sessions', f'{telegram_account.phone_number}.session'),
        ]
        
        for session_file in session_files:
            if os.path.exists(session_file):
                try:
                    os.remove(session_file)
                except Exception as e:
                    print(f"Error deleting session file {session_file}: {e}")
        
        # Delete from database (cascade will delete related chats and messages)
        phone = telegram_account.phone_number
        telegram_account.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Account {phone} butunlay o\'chirildi!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@require_http_methods(["POST"])
def update_telegram_profile(request, account_id):
    """Update Telegram profile information"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    try:
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        username = request.POST.get('username', '').strip()
        bio = request.POST.get('bio', '').strip()
        
        async def update_profile_async():
            try:
                manager = TelegramManager(telegram_account)
                await manager.connect()
                try:
                    if not await manager.is_authorized():
                        return {'success': False, 'error': 'Account avtorizatsiya qilinmagan'}
                    
                    from telethon.tl.functions.account import UpdateProfileRequest, UpdateUsernameRequest
                    from telethon.tl.functions.users import GetFullUserRequest
                    
                    # Get current profile info
                    me = await manager.client.get_me()
                    full = await manager.client(GetFullUserRequest(me.id))
                    
                    current_first_name = me.first_name or ''
                    current_last_name = me.last_name or ''
                    current_username = me.username or ''
                    current_bio = full.full_user.about or ''
                    
                    updated_fields = []
                    
                    # Check if name or bio changed
                    if (first_name != current_first_name or 
                        last_name != current_last_name or 
                        bio != current_bio):
                        await manager.client(UpdateProfileRequest(
                            first_name=first_name or 'User',
                            last_name=last_name or '',
                            about=bio or ''
                        ))
                        if first_name != current_first_name:
                            updated_fields.append('Ism')
                        if last_name != current_last_name:
                            updated_fields.append('Familiya')
                        if bio != current_bio:
                            updated_fields.append('Bio')
                    
                    # Check if username changed
                    if username != current_username:
                        try:
                            await manager.client(UpdateUsernameRequest(username))
                            updated_fields.append('Username')
                        except Exception as e:
                            return {'success': False, 'error': f'Username o\'zgartirish xatoligi: {str(e)}'}
                    
                    if updated_fields:
                        message = f'Yangilandi: {", ".join(updated_fields)}'
                    else:
                        message = 'Hech qanday o\'zgarish kiritilmadi'
                    
                    return {
                        'success': True,
                        'message': message
                    }
                    
                finally:
                    await manager.disconnect()
                    
            except Exception as e:
                return {'success': False, 'error': str(e)}
        
        result_wrapper = run_coroutine_in_thread(update_profile_async(), timeout=30)
        
        if result_wrapper.get('success'):
            result = result_wrapper.get('result', {})
            if result.get('success'):
                return JsonResponse({
                    'success': True,
                    'message': result.get('message')
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': result.get('error', 'Noma\'lum xatolik')
                })
        else:
            return JsonResponse({
                'success': False,
                'error': result_wrapper.get('error', 'Noma\'lum xatolik')
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@require_http_methods(["POST"])
def delete_all_chats(request, account_id):
    """Delete all chats/dialogs from Telegram account"""
    telegram_account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
    
    try:
        async def delete_chats_async():
            try:
                manager = TelegramManager(telegram_account)
                await manager.connect()
                try:
                    if not await manager.is_authorized():
                        return {'success': False, 'error': 'Account avtorizatsiya qilinmagan'}
                    
                    from telethon.tl.functions.messages import DeleteHistoryRequest
                    from telethon.tl.functions.channels import LeaveChannelRequest
                    
                    # Get all dialogs
                    dialogs = await manager.client.get_dialogs()
                    
                    if not dialogs:
                        return {'success': True, 'message': 'Chatlar topilmadi', 'deleted_count': 0, 'left_count': 0}
                    
                    deleted_count = 0
                    left_count = 0
                    errors = []
                    
                    for dialog in dialogs:
                        entity = dialog.entity
                        chat_name = dialog.name or "Unknown"
                        
                        try:
                            # Check if it's a group/channel/supergroup
                            if getattr(entity, "megagroup", False) or getattr(entity, "broadcast", False):
                                try:
                                    # Leave from group/channel
                                    await manager.client(LeaveChannelRequest(entity))
                                    left_count += 1
                                    print(f"✅ Left group/channel: {chat_name}")
                                except Exception as e:
                                    print(f"⚠️ Could not leave {chat_name}: {e}")
                                    errors.append(f"{chat_name}: {str(e)}")
                            else:
                                # For private chats, try to delete history
                                try:
                                    await manager.client(DeleteHistoryRequest(
                                        peer=entity,
                                        max_id=0,
                                        revoke=True  # Delete from both sides
                                    ))
                                    deleted_count += 1
                                    print(f"✅ Deleted history: {chat_name}")
                                except Exception as e:
                                    print(f"⚠️ Could not delete history for {chat_name}: {e}")
                                    # Still count as processed
                                    deleted_count += 1
                            
                        except Exception as e:
                            print(f"❌ Error processing {chat_name}: {e}")
                            errors.append(f"{chat_name}: {str(e)}")
                            continue
                    
                    message = f'{deleted_count} ta chat tozalandi, {left_count} ta guruh/kanaldan chiqildi!'
                    if errors:
                        message += f' ({len(errors)} ta xatolik bo\'ldi)'
                    
                    return {
                        'success': True, 
                        'message': message,
                        'deleted_count': deleted_count,
                        'left_count': left_count,
                        'errors': errors
                    }
                    
                finally:
                    await manager.disconnect()
                    
            except Exception as e:
                return {'success': False, 'error': str(e)}
        
        # Run with proper event loop management
        result_wrapper = run_coroutine_in_thread(delete_chats_async(), timeout=120)
        
        if result_wrapper.get('success'):
            result = result_wrapper.get('result', {})
            if result.get('success'):
                return JsonResponse({
                    'success': True,
                    'message': result.get('message'),
                    'deleted_count': result.get('deleted_count', 0),
                    'left_count': result.get('left_count', 0)
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': result.get('error', 'Noma\'lum xatolik')
                })
        else:
            return JsonResponse({
                'success': False,
                'error': result_wrapper.get('error', 'Noma\'lum xatolik')
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def export_contacts(request):
    """Export contacts to Excel file"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime
    
    account_id = request.GET.get('account')
    
    # Get contacts
    contacts = Contact.objects.filter(telegram_account__user=request.user)
    
    if account_id:
        contacts = contacts.filter(telegram_account_id=account_id)
        try:
            account = TelegramAccount.objects.get(id=account_id, user=request.user)
            filename = f'contacts_{account.phone_number}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        except TelegramAccount.DoesNotExist:
            filename = f'contacts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    else:
        filename = f'all_contacts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    contacts = contacts.select_related('telegram_account').order_by('-created_at')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kontaktlar"
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        "№", "Ism", "Familiya", "To'liq ism", "Telefon", 
        "Username", "Account", "Telegram mavjud", "Qo'shilgan", 
        "Premium", "Bot", "Yaratilgan"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Set column widths
    column_widths = [5, 20, 20, 25, 18, 20, 18, 12, 12, 10, 10, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    # Data
    for row_num, contact in enumerate(contacts, 2):
        ws.cell(row=row_num, column=1).value = row_num - 1
        ws.cell(row=row_num, column=2).value = contact.first_name or ""
        ws.cell(row=row_num, column=3).value = contact.last_name or ""
        ws.cell(row=row_num, column=4).value = contact.name or ""
        ws.cell(row=row_num, column=5).value = contact.phone_number
        ws.cell(row=row_num, column=6).value = f"@{contact.username}" if contact.username else ""
        ws.cell(row=row_num, column=7).value = contact.telegram_account.phone_number if contact.telegram_account else ""
        ws.cell(row=row_num, column=8).value = "Ha" if contact.telegram_exists else "Yo'q"
        ws.cell(row=row_num, column=9).value = "Ha" if contact.added_to_telegram else "Yo'q"
        ws.cell(row=row_num, column=10).value = "Ha" if contact.is_premium else ""
        ws.cell(row=row_num, column=11).value = "Ha" if contact.is_bot else ""
        ws.cell(row=row_num, column=12).value = contact.created_at.strftime('%Y-%m-%d %H:%M:%S')
        
        # Center align some columns
        for col in [1, 8, 9, 10, 11]:
            ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="center")
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


# ======================== AUTO-REPLY VIEWS ========================

@login_required
def list_auto_reply_rules(request, account_id):
    """List all auto-reply rules for account"""
    try:
        account = TelegramAccount.objects.get(id=account_id, user=request.user)
        rules = AutoReplyRule.objects.filter(telegram_account=account).order_by('-is_active', '-created_at')
        
        rules_data = []
        for rule in rules:
            # Convert keywords from text field (newline separated) to list
            keywords_list = rule.get_keywords_list()
            # Convert excluded_users from text field to list
            excluded_users_list = [u.strip() for u in rule.excluded_users.split('\n') if u.strip()] if rule.excluded_users else []
            
            rules_data.append({
                'id': rule.id,
                'name': rule.name,
                'is_active': rule.is_active,
                'trigger_type': rule.trigger_type,
                'trigger_type_display': rule.get_trigger_type_display(),
                'keywords': keywords_list,
                'reply_message': rule.reply_message,
                'delay_seconds': rule.delay_seconds,
                'reply_once_per_user': rule.reply_once_per_user,
                'work_hours_only': rule.work_hours_only,
                'work_hours_start': str(rule.work_hours_start) if rule.work_hours_start else None,
                'work_hours_end': str(rule.work_hours_end) if rule.work_hours_end else None,
                'excluded_users': excluded_users_list,
                'mark_as_read': rule.mark_as_read,
                'show_typing': rule.show_typing,
                'typing_duration': rule.typing_duration,
                'usage_count': rule.messages_sent_count,
                'created_at': rule.created_at.isoformat()
            })
        
        return JsonResponse({'success': True, 'rules': rules_data})
    except TelegramAccount.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Account topilmadi'}, status=404)
    except Exception as e:
        logger.error(f"Error listing auto-reply rules: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def create_auto_reply_rule(request, account_id):
    """Create new auto-reply rule"""
    try:
        account = TelegramAccount.objects.get(id=account_id, user=request.user)
        
        data = json.loads(request.body)
        
        # Validate required fields
        required_fields = ['name', 'trigger_type', 'reply_message']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({'success': False, 'error': f'{field} majburiy'}, status=400)
        
        # Convert keywords list to newline-separated text
        keywords_text = '\n'.join(data.get('keywords', []))
        # Convert excluded_users list to newline-separated text
        excluded_users_text = '\n'.join([str(u) for u in data.get('excluded_users', [])])
        
        # Create rule
        rule = AutoReplyRule.objects.create(
            telegram_account=account,
            name=data['name'],
            trigger_type=data['trigger_type'],
            keywords=keywords_text,
            reply_message=data['reply_message'],
            delay_seconds=data.get('delay_seconds', 0),
            reply_once_per_user=data.get('reply_once_per_user', False),
            work_hours_only=data.get('work_hours_only', False),
            work_hours_start=data.get('work_hours_start'),
            work_hours_end=data.get('work_hours_end'),
            excluded_users=excluded_users_text,
            only_private_chats=data.get('only_private_chats', True),
            mark_as_read=data.get('mark_as_read', True),
            show_typing=data.get('show_typing', True),
            typing_duration=data.get('typing_duration', 2),
            is_active=data.get('is_active', True)
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'Auto-reply qoidasi yaratildi',
            'rule_id': rule.id
        })
    except TelegramAccount.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Account topilmadi'}, status=404)
    except Exception as e:
        logger.error(f"Error creating auto-reply rule: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def update_auto_reply_rule(request, rule_id):
    """Update existing auto-reply rule"""
    try:
        rule = AutoReplyRule.objects.get(
            id=rule_id, 
            telegram_account__user=request.user
        )
        
        data = json.loads(request.body)
        
        # Update fields
        if 'name' in data:
            rule.name = data['name']
        if 'trigger_type' in data:
            rule.trigger_type = data['trigger_type']
        if 'keywords' in data:
            rule.keywords = '\n'.join(data['keywords'])
        if 'reply_message' in data:
            rule.reply_message = data['reply_message']
        if 'delay_seconds' in data:
            rule.delay_seconds = data['delay_seconds']
        if 'reply_once_per_user' in data:
            rule.reply_once_per_user = data['reply_once_per_user']
        if 'work_hours_only' in data:
            rule.work_hours_only = data['work_hours_only']
        if 'work_hours_start' in data:
            rule.work_hours_start = data['work_hours_start']
        if 'work_hours_end' in data:
            rule.work_hours_end = data['work_hours_end']
        if 'excluded_users' in data:
            rule.excluded_users = '\n'.join([str(u) for u in data['excluded_users']])
        if 'only_private_chats' in data:
            rule.only_private_chats = data['only_private_chats']
        if 'mark_as_read' in data:
            rule.mark_as_read = data['mark_as_read']
        if 'show_typing' in data:
            rule.show_typing = data['show_typing']
        if 'typing_duration' in data:
            rule.typing_duration = data['typing_duration']
        if 'is_active' in data:
            rule.is_active = data['is_active']
        
        rule.save()
        
        return JsonResponse({'success': True, 'message': 'Auto-reply qoidasi yangilandi'})
    except AutoReplyRule.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Qoida topilmadi'}, status=404)
    except Exception as e:
        logger.error(f"Error updating auto-reply rule: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def delete_auto_reply_rule(request, rule_id):
    """Delete auto-reply rule"""
    try:
        rule = AutoReplyRule.objects.get(
            id=rule_id,
            telegram_account__user=request.user
        )
        
        rule_name = rule.name
        rule.delete()
        
        return JsonResponse({'success': True, 'message': f'"{rule_name}" qoidasi o\'chirildi'})
    except AutoReplyRule.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Qoida topilmadi'}, status=404)
    except Exception as e:
        logger.error(f"Error deleting auto-reply rule: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def toggle_auto_reply_rule(request, rule_id):
    """Toggle auto-reply rule active status"""
    try:
        rule = AutoReplyRule.objects.get(
            id=rule_id,
            telegram_account__user=request.user
        )
        
        rule.is_active = not rule.is_active
        rule.save()
        
        status = 'aktivlashtirildi' if rule.is_active else 'o\'chirildi'
        return JsonResponse({
            'success': True, 
            'message': f'Auto-reply {status}',
            'is_active': rule.is_active
        })
    except AutoReplyRule.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Qoida topilmadi'}, status=404)
    except Exception as e:
        logger.error(f"Error toggling auto-reply rule: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def start_monitoring_view(request):
    """Start monitoring service"""
    try:
        from .telegram_monitor import start_monitoring
        start_monitoring()
        return JsonResponse({'success': True, 'message': 'Monitoring ishga tushirildi'})
    except Exception as e:
        logger.error(f"Error starting monitoring: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def monitoring_status(request):
    """Get monitoring status"""
    try:
        from .telegram_monitor import get_monitor
        monitor = get_monitor()
        return JsonResponse({
            'success': True,
            'running': monitor.running,
            'check_interval': monitor.check_interval
        })
    except Exception as e:
        logger.error(f"Error getting monitoring status: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================
# AI ASSISTANT VIEWS
# ============================================

@login_required
def ai_dashboard(request):
    """AI dashboard page"""
    from .models import AIProvider
    providers = AIProvider.objects.filter(user=request.user)
    accounts = TelegramAccount.objects.filter(user=request.user)
    
    context = {
        'providers': providers,
        'accounts': accounts,
    }
    return render(request, 'home/ai_dashboard.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def ai_provider_manage(request, provider_id=None):
    """Create or update AI provider"""
    from .models import AIProvider
    
    if request.method == 'GET':
        if provider_id:
            # Edit existing provider
            provider = get_object_or_404(AIProvider, id=provider_id, user=request.user)
            return JsonResponse({
                'success': True,
                'provider': {
                    'id': provider.id,
                    'name': provider.name,
                    'provider_type': provider.provider_type,
                    'api_key': provider.api_key,
                    'api_endpoint': provider.api_endpoint or '',
                    'is_active': provider.is_active,
                }
            })
        else:
            # List all providers
            providers = AIProvider.objects.filter(user=request.user)
            return JsonResponse({
                'success': True,
                'providers': [{
                    'id': p.id,
                    'name': p.name,
                    'provider_type': p.provider_type,
                    'provider_type_display': p.get_provider_type_display(),
                    'is_active': p.is_active,
                    'created_at': p.created_at.isoformat(),
                } for p in providers]
            })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            if provider_id:
                # Update existing
                provider = get_object_or_404(AIProvider, id=provider_id, user=request.user)
                provider.name = data.get('name', provider.name)
                provider.provider_type = data.get('provider_type', provider.provider_type)
                provider.api_key = data.get('api_key', provider.api_key)
                provider.api_endpoint = data.get('api_endpoint', provider.api_endpoint)
                provider.is_active = data.get('is_active', provider.is_active)
                provider.save()
                message = 'AI Provider yangilandi'
            else:
                # Create new
                provider = AIProvider.objects.create(
                    user=request.user,
                    name=data['name'],
                    provider_type=data['provider_type'],
                    api_key=data['api_key'],
                    api_endpoint=data.get('api_endpoint', ''),
                    is_active=data.get('is_active', True)
                )
                message = 'AI Provider yaratildi'
            
            return JsonResponse({
                'success': True,
                'message': message,
                'provider_id': provider.id
            })
        except Exception as e:
            logger.error(f"Error managing AI provider: {e}", exc_info=True)
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ai_provider_delete(request, provider_id):
    """Delete AI provider"""
    from .models import AIProvider
    try:
        provider = get_object_or_404(AIProvider, id=provider_id, user=request.user)
        provider_name = provider.name
        provider.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'"{provider_name}" o\'chirildi'
        })
    except Exception as e:
        logger.error(f"Error deleting AI provider: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET", "POST"])
def ai_assistant_manage(request, assistant_id=None):
    """Create or update AI assistant"""
    from .models import AIAssistant, AIProvider
    
    if request.method == 'GET':
        if assistant_id:
            # Get specific assistant
            assistant = get_object_or_404(
                AIAssistant.objects.select_related('telegram_account', 'ai_provider'),
                id=assistant_id,
                telegram_account__user=request.user
            )
            return JsonResponse({
                'success': True,
                'assistant': {
                    'id': assistant.id,
                    'name': assistant.name,
                    'telegram_account_id': assistant.telegram_account.id,
                    'ai_provider_id': assistant.ai_provider.id,
                    'model': assistant.model,
                    'system_prompt': assistant.system_prompt,
                    'is_active': assistant.is_active,
                    'auto_respond': assistant.auto_respond,
                    'only_private_chats': assistant.only_private_chats,
                    'response_delay_seconds': assistant.response_delay_seconds,
                    'mark_as_read': assistant.mark_as_read,
                    'show_typing': assistant.show_typing,
                    'typing_duration': assistant.typing_duration,
                }
            })
        else:
            # List all assistants
            assistants = AIAssistant.objects.filter(
                telegram_account__user=request.user
            ).select_related('telegram_account', 'ai_provider')
            
            return JsonResponse({
                'success': True,
                'assistants': [{
                    'id': a.id,
                    'name': a.name,
                    'telegram_account': a.telegram_account.phone_number,
                    'telegram_account_id': a.telegram_account.id,
                    'ai_provider': a.ai_provider.name,
                    'ai_provider_id': a.ai_provider.id,
                    'model': a.model,
                    'model_display': a.get_model_display(),
                    'is_active': a.is_active,
                    'auto_respond': a.auto_respond,
                    'messages_processed': a.messages_processed,
                    'created_at': a.created_at.isoformat(),
                } for a in assistants]
            })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Validate required fields
            if not assistant_id:
                required_fields = ['name', 'telegram_account_id', 'ai_provider_id', 'model']
                for field in required_fields:
                    if not data.get(field):
                        return JsonResponse({
                            'success': False,
                            'error': f'{field} majburiy'
                        }, status=400)
            
            if assistant_id:
                # Update existing
                assistant = get_object_or_404(
                    AIAssistant,
                    id=assistant_id,
                    telegram_account__user=request.user
                )
                assistant.name = data.get('name', assistant.name)
                assistant.model = data.get('model', assistant.model)
                assistant.system_prompt = data.get('system_prompt', assistant.system_prompt)
                assistant.is_active = data.get('is_active', assistant.is_active)
                assistant.auto_respond = data.get('auto_respond', assistant.auto_respond)
                assistant.only_private_chats = data.get('only_private_chats', assistant.only_private_chats)
                assistant.response_delay_seconds = data.get('response_delay_seconds', assistant.response_delay_seconds)
                assistant.mark_as_read = data.get('mark_as_read', assistant.mark_as_read)
                assistant.show_typing = data.get('show_typing', assistant.show_typing)
                assistant.typing_duration = data.get('typing_duration', assistant.typing_duration)
                
                if 'ai_provider_id' in data:
                    provider = get_object_or_404(AIProvider, id=data['ai_provider_id'], user=request.user)
                    assistant.ai_provider = provider
                
                assistant.save()
                message = 'AI Assistant yangilandi'
            else:
                # Create new
                account = get_object_or_404(
                    TelegramAccount,
                    id=data['telegram_account_id'],
                    user=request.user
                )
                provider = get_object_or_404(
                    AIProvider,
                    id=data['ai_provider_id'],
                    user=request.user
                )
                
                assistant = AIAssistant.objects.create(
                    telegram_account=account,
                    ai_provider=provider,
                    name=data['name'],
                    model=data['model'],
                    system_prompt=data.get('system_prompt', ''),
                    is_active=data.get('is_active', True),
                    auto_respond=data.get('auto_respond', True),
                    only_private_chats=data.get('only_private_chats', True),
                    response_delay_seconds=data.get('response_delay_seconds', 2),
                    mark_as_read=data.get('mark_as_read', True),
                    show_typing=data.get('show_typing', True),
                    typing_duration=data.get('typing_duration', 3)
                )
                message = 'AI Assistant yaratildi'
            
            return JsonResponse({
                'success': True,
                'message': message,
                'assistant_id': assistant.id
            })
        except Exception as e:
            logger.error(f"Error managing AI assistant: {e}", exc_info=True)
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ai_assistant_delete(request, assistant_id):
    """Delete AI assistant"""
    from .models import AIAssistant
    try:
        assistant = get_object_or_404(
            AIAssistant,
            id=assistant_id,
            telegram_account__user=request.user
        )
        assistant_name = assistant.name
        assistant.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'"{assistant_name}" o\'chirildi'
        })
    except Exception as e:
        logger.error(f"Error deleting AI assistant: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def ai_assistant_toggle(request, assistant_id):
    """Toggle AI assistant active status"""
    from .models import AIAssistant
    try:
        assistant = get_object_or_404(
            AIAssistant,
            id=assistant_id,
            telegram_account__user=request.user
        )
        assistant.is_active = not assistant.is_active
        assistant.save()
        
        return JsonResponse({
            'success': True,
            'is_active': assistant.is_active,
            'message': f'AI Assistant {"yoqildi" if assistant.is_active else "o\'chirildi"}'
        })
    except Exception as e:
        logger.error(f"Error toggling AI assistant: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def ai_conversation_summaries(request, account_id):
    """Get conversation summaries for account"""
    try:
        from .models import ConversationSummary
        
        account = get_object_or_404(TelegramAccount, id=account_id, user=request.user)
        summaries = ConversationSummary.objects.filter(
            telegram_account=account
        ).select_related('ai_assistant').order_by('-last_message_at')
        
        return JsonResponse({
            'success': True,
            'summaries': [{
                'id': s.id,
                'chat_id': s.chat_id,
                'user_id': s.user_id,
                'username': s.username,
                'ai_assistant': s.ai_assistant.name,
                'summary_data': s.summary_data,
                'last_message_at': s.last_message_at.isoformat(),
            } for s in summaries]
        })
    except Exception as e:
        logger.error(f"Error getting conversation summaries: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ==================== CRM Integration Views ====================

@login_required
def crm_integration_dashboard(request):
    """CRM Integration dashboard"""
    from .models import CRMProvider, PropertySearchLog
    
    providers = CRMProvider.objects.filter(user=request.user)
    recent_searches = PropertySearchLog.objects.filter(
        crm_provider__user=request.user
    )[:20]
    
    context = {
        'crm_providers': providers,
        'recent_searches': recent_searches,
    }
    return render(request, 'home/crm_dashboard.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def crm_provider_manage(request, provider_id=None):
    """Create or update CRM provider"""
    from .models import CRMProvider
    
    if request.method == 'GET':
        if provider_id:
            # Get single provider
            provider = get_object_or_404(CRMProvider, id=provider_id, user=request.user)
            return JsonResponse({
                'success': True,
                'provider': {
                    'id': provider.id,
                    'name': provider.name,
                    'crm_type': provider.crm_type,
                    'api_url': provider.api_url or '',
                    'api_key': provider.api_key or '',
                    'api_secret': provider.api_secret or '',
                    'field_mapping': provider.field_mapping,
                    'request_template': provider.request_template,
                    'extraction_prompt': provider.extraction_prompt or '',
                    'is_active': provider.is_active,
                }
            })
        else:
            # List all providers
            providers = CRMProvider.objects.filter(user=request.user)
            return JsonResponse({
                'success': True,
                'providers': [{
                    'id': p.id,
                    'name': p.name,
                    'crm_type': p.crm_type,
                    'crm_type_display': p.get_crm_type_display(),
                    'api_url': p.api_url,
                    'is_active': p.is_active,
                    'created_at': p.created_at.isoformat(),
                } for p in providers]
            })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            if provider_id:
                # Update existing
                provider = get_object_or_404(CRMProvider, id=provider_id, user=request.user)
                provider.name = data.get('name', provider.name)
                provider.crm_type = data.get('crm_type', provider.crm_type)
                provider.api_url = data.get('api_url', provider.api_url)
                provider.api_key = data.get('api_key', provider.api_key)
                provider.api_secret = data.get('api_secret', provider.api_secret)
                provider.field_mapping = data.get('field_mapping', provider.field_mapping)
                provider.request_template = data.get('request_template', provider.request_template)
                provider.extraction_prompt = data.get('extraction_prompt', provider.extraction_prompt)
                provider.is_active = data.get('is_active', provider.is_active)
                provider.save()
                message = 'CRM Provider yangilandi'
            else:
                # Create new
                provider = CRMProvider.objects.create(
                    user=request.user,
                    name=data['name'],
                    crm_type=data['crm_type'],
                    api_url=data.get('api_url', ''),
                    api_key=data.get('api_key', ''),
                    api_secret=data.get('api_secret', ''),
                    field_mapping=data.get('field_mapping', {}),
                    request_template=data.get('request_template', {}),
                    extraction_prompt=data.get('extraction_prompt', ''),
                    is_active=data.get('is_active', True)
                )
                message = 'CRM Provider yaratildi'
            
            return JsonResponse({
                'success': True,
                'message': message,
                'provider_id': provider.id
            })
            
        except Exception as e:
            logger.error(f"Error managing CRM provider: {e}", exc_info=True)
            return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def crm_provider_delete(request, provider_id):
    """Delete CRM provider"""
    try:
        from .models import CRMProvider
        provider = get_object_or_404(CRMProvider, id=provider_id, user=request.user)
        provider_name = provider.name
        provider.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'"{provider_name}" o\'chirildi'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def crm_test_connection(request, provider_id):
    """Test CRM connection"""
    try:
        from .models import CRMProvider
        from .crm_service import CRMService
        
        provider = get_object_or_404(CRMProvider, id=provider_id, user=request.user)
        crm_service = CRMService(provider)
        
        # Run async test
        result = run_async(crm_service.test_connection())
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f"CRM test error: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def crm_search_properties(request):
    """Search properties in CRM based on conversation"""
    try:
        from .models import CRMProvider, ConversationSummary, PropertySearchLog, AIProvider
        from .crm_service import CRMService
        from .ai_service import AIService
        
        data = json.loads(request.body)
        crm_provider_id = data.get('crm_provider_id')
        conversation_summary_id = data.get('conversation_summary_id')
        
        # Get objects
        crm_provider = get_object_or_404(CRMProvider, id=crm_provider_id, user=request.user)
        conversation = get_object_or_404(
            ConversationSummary, 
            id=conversation_summary_id,
            telegram_account__user=request.user
        )
        
        # Get AI provider for extraction
        ai_assistant = conversation.ai_assistant
        ai_provider_obj = ai_assistant.ai_provider
        
        # Initialize services
        ai_service = AIService(
            provider_type=ai_provider_obj.provider_type,
            api_key=ai_provider_obj.api_key,
            model=ai_assistant.model,
            api_endpoint=ai_provider_obj.api_endpoint
        )
        crm_service = CRMService(crm_provider)
        
        # Step 1: Extract requirements using AI
        extraction_result = run_async(
            crm_service.extract_requirements_with_ai(
                conversation.summary_data,
                ai_service
            )
        )
        
        if not extraction_result.get('success'):
            return JsonResponse({
                'success': False,
                'error': 'AI tahlil xatolik: ' + extraction_result.get('error', 'Unknown')
            })
        
        requirements = extraction_result['requirements']
        
        # Step 2: Search in CRM
        search_result = run_async(
            crm_service.search_properties(requirements)
        )
        
        # Step 3: Save log
        log = PropertySearchLog.objects.create(
            crm_provider=crm_provider,
            telegram_account=conversation.telegram_account,
            chat_id=conversation.chat_id,
            username=conversation.username,
            extracted_requirements=requirements,
            crm_request=search_result.get('raw_response', {}),
            crm_response=search_result.get('raw_response', {}),
            results_count=search_result.get('count', 0),
            status='success' if search_result.get('success') else 'failed',
            error_message=search_result.get('error') if not search_result.get('success') else None
        )
        
        return JsonResponse({
            'success': search_result.get('success'),
            'requirements': requirements,
            'properties': search_result.get('properties', []),
            'count': search_result.get('count', 0),
            'log_id': log.id
        })
        
    except Exception as e:
        logger.error(f"CRM search error: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
def property_interests_dashboard(request):
    """
    Takliflar dashboard - mijozlar qiziqish bildirgan uylar
    """
    from .models import PropertyInterest
    
    # Get filters
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('q', '')
    
    # Base queryset
    interests = PropertyInterest.objects.filter(
        telegram_account__user=request.user
    ).select_related('telegram_account', 'contact', 'search_log').order_by('-created_at')
    
    # Apply filters
    if status_filter:
        interests = interests.filter(status=status_filter)
    
    if search_query:
        interests = interests.filter(
            Q(username__icontains=search_query) |
            Q(property_id__icontains=search_query) |
            Q(contact__phone_number__icontains=search_query)
        )
    
    # Stats
    total_interests = interests.count()
    interested_count = interests.filter(status='interested').count()
    rejected_count = interests.filter(status='rejected').count()
    viewed_count = interests.filter(status='viewed').count()
    
    # Pagination
    paginator = Paginator(interests, 50)
    page = request.GET.get('page', 1)
    interests_page = paginator.get_page(page)
    
    context = {
        'interests': interests_page,
        'total_interests': total_interests,
        'interested_count': interested_count,
        'rejected_count': rejected_count,
        'viewed_count': viewed_count,
        'status_filter': status_filter,
        'search_query': search_query,
    }
    return render(request, 'home/property_interests.html', context)


@login_required
@require_http_methods(["GET"])
def property_interest_detail(request, interest_id):
    """Get property interest details"""
    from .models import PropertyInterest
    
    try:
        interest = get_object_or_404(
            PropertyInterest,
            id=interest_id,
            telegram_account__user=request.user
        )
        
        return JsonResponse({
            'success': True,
            'interest': {
                'id': interest.id,
                'property_id': interest.property_id,
                'property_data': interest.property_data,
                'username': interest.username,
                'chat_id': interest.chat_id,
                'contact': {
                    'name': interest.contact.name if interest.contact else None,
                    'phone': interest.contact.phone_number if interest.contact else None,
                } if interest.contact else None,
                'status': interest.status,
                'created_at': interest.created_at.isoformat(),
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

