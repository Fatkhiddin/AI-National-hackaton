# market_analysis/management/commands/import_from_excel.py

"""
Excel fayllardan bozor narxlarini import qilish.
3 ta Excel fayldan narxlarning o'rtachasini olib import qiladi.
"""

from django.core.management.base import BaseCommand
from django.db import transaction
import openpyxl
import os
from decimal import Decimal
from market_analys.models import MarketPriceReference
from users.models import Team


class Command(BaseCommand):
    help = 'Excel fayllardan bozor narxlarini import qilish (3 ta faylning o\'rtachasi)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--team',
            type=str,
            help='Team domain nomi (default: crm.megapolis1.uz)',
            default='crm.megapolis1.uz'
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Avval barcha mavjud ma\'lumotlarni o\'chirish'
        )

    def handle(self, *args, **options):
        team_domain = options['team']
        clear = options.get('clear', False)

        # Team tekshirish (domain bo'yicha)
        try:
            team = Team.objects.get(domain=team_domain)
        except Team.DoesNotExist:
            self.stdout.write(self.style.ERROR(f'❌ Team domain "{team_domain}" topilmadi'))
            return

        self.stdout.write(self.style.SUCCESS(f'\n{"="*60}'))
        self.stdout.write(self.style.SUCCESS(f'📊 EXCEL IMPORT - TEAM: {team.name}'))
        self.stdout.write(self.style.SUCCESS(f'{"="*60}\n'))

        # Agar --clear bo'lsa, avval o'chirish
        if clear:
            count = MarketPriceReference.objects.filter(team=team).count()
            MarketPriceReference.objects.filter(team=team).delete()
            self.stdout.write(self.style.WARNING(f'🗑️  {count} ta eski yozuv o\'chirildi\n'))

        # Excel fayllar yo'li
        excel_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
            'excels'
        )
        
        excel_files = [
            os.path.join(excel_dir, 'Price 1.xlsx'),
            os.path.join(excel_dir, 'Price 2.xlsx'),
            os.path.join(excel_dir, 'Price 3.xlsx'),
        ]

        # Fayllar borligini tekshirish
        for file_path in excel_files:
            if not os.path.exists(file_path):
                self.stdout.write(self.style.ERROR(f'❌ Fayl topilmadi: {file_path}'))
                return

        self.stdout.write(self.style.SUCCESS('✅ 3 ta Excel fayl topildi\n'))

        # 3 ta fayldan ma'lumotlarni o'qish
        all_data = {}  # key: (etaj, xonalar, qurilish, holat, maydon), value: [narxlar_list]

        for idx, file_path in enumerate(excel_files, 1):
            self.stdout.write(f'📖 {idx}. {os.path.basename(file_path)} o\'qilmoqda...')
            
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                
                # Remontli va Remontsiz sheetlar
                for sheet_name in ['Remontli', 'Remontsiz']:
                    if sheet_name not in wb.sheetnames:
                        continue
                    
                    sheet = wb[sheet_name]
                    holat = 'remontli' if sheet_name == 'Remontli' else 'remontsiz'
                    
                    # Qatorlarni o'qish (3-qatordan boshlab, chunki 1-2 header)
                    # Avval barcha qatorlarni o'qib olamiz
                    rows_data = []
                    
                    for row_idx in range(3, sheet.max_row + 1):
                        try:
                            etaj = sheet.cell(row_idx, 2).value  # B ustun
                            xonalar = sheet.cell(row_idx, 3).value  # C ustun
                            qurilish = sheet.cell(row_idx, 4).value  # D ustun
                            maydon = sheet.cell(row_idx, 5).value  # E ustun
                            arzon = sheet.cell(row_idx, 6).value  # F ustun
                            bozor = sheet.cell(row_idx, 7).value  # G ustun
                            qimmat = sheet.cell(row_idx, 8).value  # H ustun

                            # Bo'sh qatorlarni o'tkazish
                            if not etaj or not xonalar or not qurilish:
                                continue

                            # Maydon bo'sh bo'lsa o'tkazish
                            if not maydon:
                                continue

                            # Narxlar 0 bo'lsa o'tkazish (uy mavjud emas)
                            if not arzon or not bozor or not qimmat:
                                continue
                            if arzon == 0 or bozor == 0 or qimmat == 0:
                                continue

                            rows_data.append({
                                'etaj': etaj,
                                'xonalar': xonalar,
                                'qurilish': qurilish,
                                'maydon': maydon,
                                'arzon': arzon,
                                'bozor': bozor,
                                'qimmat': qimmat,
                            })

                        except Exception as e:
                            continue

                    # Endi maydon maksimal va minimallarni hisoblash
                    for i, row_data in enumerate(rows_data):
                        try:
                            etaj = row_data['etaj']
                            xonalar = row_data['xonalar']
                            qurilish = row_data['qurilish']
                            maydon = row_data['maydon']
                            arzon = row_data['arzon']
                            bozor = row_data['bozor']
                            qimmat = row_data['qimmat']

                            # Maydon minimal - hozirgi maydon
                            # Maydon maksimal - keyingi qatordagi maydon (agar mavjud bo'lsa)
                            maydon_min = int(maydon)
                            maydon_max = None
                            
                            if i + 1 < len(rows_data):
                                # Keyingi qatordagi maydon
                                next_maydon = rows_data[i + 1]['maydon']
                                maydon_max = int(next_maydon)

                            # Qurilish turini normalize qilish
                            qurilish_lower = str(qurilish).lower().strip()
                            if 'гишт' in qurilish_lower or 'кирпич' in qurilish_lower or 'g\'isht' in qurilish_lower or 'gishtli' in qurilish_lower:
                                qurilish_normalized = 'gishtli'
                            elif 'панел' in qurilish_lower or 'panel' in qurilish_lower:
                                qurilish_normalized = 'panelli'
                            elif 'монолит' in qurilish_lower or 'monolit' in qurilish_lower:
                                qurilish_normalized = 'monolitli'
                            elif 'блок' in qurilish_lower or 'blok' in qurilish_lower:
                                qurilish_normalized = 'blokli'
                            else:
                                qurilish_normalized = 'gishtli'  # default

                            # Key yaratish
                            key = (int(etaj), int(xonalar), qurilish_normalized, holat, maydon_min, maydon_max)

                            # Ma'lumotlarni saqlash
                            if key not in all_data:
                                all_data[key] = []

                            all_data[key].append({
                                'arzon': float(arzon),
                                'bozor': float(bozor),
                                'qimmat': float(qimmat),
                            })

                        except Exception as e:
                            continue

                self.stdout.write(self.style.SUCCESS(f'   ✅ {sheet_name} o\'qildi'))

            except Exception as e:
                self.stdout.write(self.style.ERROR(f'   ❌ Xatolik: {str(e)}'))
                continue

        self.stdout.write(self.style.SUCCESS(f'\n✅ Jami {len(all_data)} ta unique yozuv topildi'))

        # O'rtachalarni hisoblash va import qilish
        success_count = 0
        error_count = 0

        self.stdout.write(self.style.SUCCESS('\n💾 Bazaga import qilinmoqda...\n'))

        with transaction.atomic():
            for key, narxlar_list in all_data.items():
                try:
                    etaj, xonalar, qurilish, holat, maydon_min, maydon_max = key

                    # O'rtachalarni hisoblash (UMUMIY UY NARXI USD)
                    arzon_avg = sum(n['arzon'] for n in narxlar_list) / len(narxlar_list)
                    bozor_avg = sum(n['bozor'] for n in narxlar_list) / len(narxlar_list)
                    qimmat_avg = sum(n['qimmat'] for n in narxlar_list) / len(narxlar_list)
                    
                    # O'rtacha maydonni hisoblash
                    if maydon_max:
                        avg_maydon = (maydon_min + maydon_max) / 2
                    else:
                        avg_maydon = maydon_min
                    
                    # 1 m² narxini hisoblash (USD)
                    # Excel da UMUMIY uy narxi bo'lgani uchun maydonga bo'lamiz
                    arzon_per_m2 = arzon_avg / avg_maydon
                    bozor_per_m2 = bozor_avg / avg_maydon
                    qimmat_per_m2 = qimmat_avg / avg_maydon

                    # Bazaga saqlash (1 m² uchun USD)
                    MarketPriceReference.objects.update_or_create(
                        team=team,
                        etaj=etaj,
                        xonalar_soni=xonalar,
                        qurilish_turi=qurilish,
                        holat=holat,
                        maydon_min=maydon_min,
                        defaults={
                            'maydon_max': maydon_max,
                            'arzon_narx': Decimal(str(round(arzon_per_m2, 2))),
                            'bozor_narx': Decimal(str(round(bozor_per_m2, 2))),
                            'qimmat_narx': Decimal(str(round(qimmat_per_m2, 2))),
                            'source_file': 'Excel Import (Price 1,2,3 average) - 1m² USD',
                        }
                    )

                    success_count += 1

                    # Progress ko'rsatish
                    if success_count % 50 == 0:
                        self.stdout.write(f'   ... {success_count} ta saqlandi')

                except Exception as e:
                    error_count += 1
                    self.stdout.write(self.style.ERROR(f'   ❌ Xatolik: {str(e)}'))

        # Natija
        self.stdout.write(self.style.SUCCESS(f'\n{"="*60}'))
        self.stdout.write(self.style.SUCCESS(f'✅ IMPORT TUGADI'))
        self.stdout.write(self.style.SUCCESS(f'{"="*60}'))
        self.stdout.write(self.style.SUCCESS(f'   Muvaffaqiyatli: {success_count}'))
        if error_count > 0:
            self.stdout.write(self.style.WARNING(f'   Xatoliklar: {error_count}'))
        self.stdout.write(self.style.SUCCESS(f'{"="*60}\n'))
